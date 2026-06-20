# database.py — HSHOP V1.0 PLATINIUM
# BIOS Creations — 2026
# Base de données SQLite complète — toutes tables + migrations

import sqlite3
import os
import logging

log = logging.getLogger("hshop")

DB_NAME = "hshop_v21.db"

# ==========================================
# FONCTIONS DE L'API WEB (AJOUTÉES EN HAUT)
# ==========================================

def lire_produits():
    """Récupère tous les produits actifs de la base HSHOP PLATINIUM."""
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT COALESCE(code_barre, code, '') AS code,
                   nom,
                   categorie,
                   prix_vente,
                   COALESCE(stock, stock_actuel, 0) AS stock_reel,
                   COALESCE(seuil_alerte, stock_min, 0) AS seuil
            FROM produits
            WHERE actif = 1
            ORDER BY nom
        """)
        lignes = cursor.fetchall()
        return [dict(ligne) for ligne in lignes]
    except sqlite3.Error as e:
        log.error(f"Erreur SQL HSHOP Produits : {e}")
        return []
    finally:
        conn.close()

def ajouter_produit(code, nom, categorie, prix_vente, stock, seuil):
    """Insère un produit selon le schéma officiel HSHOP."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO produits 
            (code_barre, code, nom, categorie, prix_vente, stock, stock_actuel, seuil_alerte, stock_min, actif)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        """, (code, code, nom, categorie, float(prix_vente), float(stock), float(stock), float(seuil), float(seuil)))
        conn.commit()
        return True
    except sqlite3.Error as e:
        log.error(f"Erreur d'insertion HSHOP : {e}")
        return False
    finally:
        conn.close()

def lire_categories():
    """Récupère toutes les catégories de la base HSHOP."""
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, nom FROM categories ORDER BY nom")
        lignes = cursor.fetchall()
        return [dict(ligne) for ligne in lignes]
    except sqlite3.Error as e:
        log.error(f"Erreur SQL Catégories : {e}")
        return []
    finally:
        conn.close()

def ajouter_categorie(nom):
    """Insère un nouveau rayon en majuscules."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO categories (nom) VALUES (?)", (nom.upper(),))
        conn.commit()
        return True
    except sqlite3.Error as e:
        log.error(f"Erreur insertion Catégorie : {e}")
        return False
    finally:
        conn.close()

# ==========================================
# FIN DES FONCTIONS API — LE RESTE DU CODE SUIT
# ==========================================

def get_db_path() -> str:
    """
    Retourne le chemin absolu vers la base de données.

    Délègue à config.get_db_path() pour avoir UNE SEULE source de vérité
    (gère le chemin défini dans app_config.json, avec retombée automatique
    sur le fichier local si le chemin configuré est absolu et invalide
    sur ce poste). Fallback local en cas d'import impossible.
    """
    try:
        from config import get_db_path as _get_db_path_config
        return _get_db_path_config()
    except Exception:
        import sys
        if getattr(sys, "frozen", False):
            base = os.path.dirname(sys.executable)
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, DB_NAME)


def get_connection() -> sqlite3.Connection:
    """
    Ouvre et retourne une connexion SQLite avec row_factory.

    NOTE DÉPLOIEMENT RÉSEAU (partage SMB serveur ↔ postes clients) :
    Le mode WAL est volontairement DÉSACTIVÉ ici. WAL repose sur de la
    mémoire partagée (fichier -shm) entre processus, ce qui ne fonctionne
    pas de manière fiable quand les processus tournent sur des machines
    différentes connectées via un partage réseau (SMB/CIFS) — c'est documenté
    explicitement par SQLite lui-même (https://sqlite.org/wal.html,
    désavantage n°2) et confirmé par de nombreux retours d'expérience
    (corruption silencieuse ou erreurs de verrou). On reste donc en mode
    DELETE (rollback journal classique), qui est conçu pour ce cas d'usage.
    Si un jour HSHOP passe en architecture client/serveur (API au lieu d'un
    partage de fichier), WAL pourra être réactivé côté serveur uniquement.
    """
    conn = sqlite3.connect(get_db_path(), timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=DELETE")  # requis pour accès réseau SMB — voir note ci-dessus
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout=10000")  # attendre 10s si DB occupée
    conn.execute("PRAGMA cache_size=-8000")     # 8MB de cache en mémoire
    conn.execute("PRAGMA synchronous=FULL")     # intégrité prioritaire en mode DELETE / réseau
    return conn

def wal_checkpoint():
    """
    Conservé pour compatibilité (appels existants dans le code).
    Sans effet réel désormais : la base n'utilise plus le mode WAL
    (voir get_connection()), donc il n'y a plus de fichier -wal à fusionner.
    Ne lève jamais d'erreur même si appelée.
    """
    try:
        conn = sqlite3.connect(get_db_path(), timeout=5)
        conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        conn.close()
        log.debug("WAL checkpoint effectué (no-op, mode DELETE actif)")
    except Exception as e:
        log.warning(f"WAL checkpoint échoué : {e}")


# ═══════════════════════════════════════════════════════════════════════════
# INITIALISATION — création de toutes les tables
# ═══════════════════════════════════════════════════════════════════════════

def init_db():
    """Crée toutes les tables si elles n'existent pas + applique les migrations."""
    conn = get_connection()
    cur  = conn.cursor()

    # ── 1. Catégories ────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            nom         TEXT    NOT NULL UNIQUE,
            description TEXT    DEFAULT '',
            created_at  TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 2. Fournisseurs ──────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS fournisseurs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            nom         TEXT    NOT NULL,
            contact     TEXT    DEFAULT '',
            telephone   TEXT    DEFAULT '',
            email       TEXT    DEFAULT '',
            adresse     TEXT    DEFAULT '',
            notes       TEXT    DEFAULT '',
            created_at  TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 3. Produits ──────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS produits (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            nom             TEXT    NOT NULL,
            code_barre      TEXT    DEFAULT '',
            categorie_id    INTEGER REFERENCES categories(id) ON DELETE SET NULL,
            fournisseur_id  INTEGER REFERENCES fournisseurs(id) ON DELETE SET NULL,
            prix_achat      REAL    DEFAULT 0.0,
            prix_vente      REAL    NOT NULL DEFAULT 0.0,
            stock           REAL    DEFAULT 0.0,
            stock_min       REAL    DEFAULT 0.0,
            unite           TEXT    DEFAULT 'pce',
            description     TEXT    DEFAULT '',
            actif           INTEGER DEFAULT 1,
            created_at      TEXT    DEFAULT (datetime('now')),
            updated_at      TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 4. Clients ───────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS clients (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            nom             TEXT    NOT NULL,
            telephone       TEXT    DEFAULT '',
            email           TEXT    DEFAULT '',
            adresse         TEXT    DEFAULT '',
            encours_max     REAL    DEFAULT 0.0,
            solde_credit    REAL    DEFAULT 0.0,
            created_at      TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 5. Utilisateurs ──────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS utilisateurs (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            identifiant  TEXT    NOT NULL UNIQUE,
            mot_de_passe TEXT    NOT NULL,
            role         TEXT    NOT NULL DEFAULT 'vendeur',
            actif        INTEGER DEFAULT 1,
            created_at   TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 6. Ventes — schéma réel (ui_caisse.py) ───────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ventes (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            date_vente       TEXT    DEFAULT (datetime('now','localtime')),
            total            REAL    NOT NULL DEFAULT 0.0,
            recu             REAL    DEFAULT 0.0,
            monnaie          REAL    DEFAULT 0.0,
            vendeur          TEXT    DEFAULT '',
            mode_paiement    TEXT    DEFAULT 'ESPECES',
            client_id        INTEGER REFERENCES clients(id) ON DELETE SET NULL,
            remise           REAL    DEFAULT 0.0,
            montant_regle    REAL    DEFAULT 0.0,
            note             TEXT    DEFAULT '',
            cloture          INTEGER DEFAULT 0,
            annulee          INTEGER DEFAULT 0,
            motif_annulation TEXT    DEFAULT ''
        )
    """)

    # ── 7. Lignes de vente ───────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS lignes_vente (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            vente_id        INTEGER NOT NULL REFERENCES ventes(id) ON DELETE CASCADE,
            produit_id      INTEGER NOT NULL REFERENCES produits(id) ON DELETE RESTRICT,
            quantite        REAL    NOT NULL DEFAULT 1.0,
            prix_unitaire   REAL    NOT NULL DEFAULT 0.0,
            remise          REAL    DEFAULT 0.0,
            sous_total      REAL    GENERATED ALWAYS AS
                                (quantite * prix_unitaire * (1 - remise/100)) STORED
        )
    """)

    # ── 8. Échéances de paiement ─────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS echeances_paiement (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            vente_id        INTEGER NOT NULL REFERENCES ventes(id) ON DELETE CASCADE,
            montant         REAL    NOT NULL DEFAULT 0.0,
            date_echeance   TEXT    NOT NULL,
            statut          TEXT    DEFAULT 'en_attente',
            date_reglement  TEXT    DEFAULT NULL,
            note            TEXT    DEFAULT '',
            created_at      TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 9. Règlements crédit ─────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS reglements_credit (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id       INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
            echeance_id     INTEGER REFERENCES echeances_paiement(id) ON DELETE SET NULL,
            montant         REAL    NOT NULL DEFAULT 0.0,
            mode_paiement   TEXT    DEFAULT 'especes',
            note            TEXT    DEFAULT '',
            date_reglement  TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 10. Commandes fournisseurs ───────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS commandes_fournisseurs (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            fournisseur_id  INTEGER REFERENCES fournisseurs(id) ON DELETE SET NULL,
            utilisateur_id  INTEGER REFERENCES utilisateurs(id) ON DELETE SET NULL,
            statut          TEXT    DEFAULT 'en_cours',
            total           REAL    DEFAULT 0.0,
            note            TEXT    DEFAULT '',
            date_commande   TEXT    DEFAULT (datetime('now')),
            date_livraison  TEXT    DEFAULT NULL
        )
    """)

    # ── 11. Lignes de commande fournisseur ───────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS lignes_commande (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            commande_id     INTEGER NOT NULL REFERENCES commandes_fournisseurs(id) ON DELETE CASCADE,
            produit_id      INTEGER NOT NULL REFERENCES produits(id) ON DELETE RESTRICT,
            quantite        REAL    NOT NULL DEFAULT 1.0,
            prix_unitaire   REAL    NOT NULL DEFAULT 0.0
        )
    """)

    # ── 12. Mouvements de stock ──────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mouvements_stock (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            produit_id      INTEGER NOT NULL REFERENCES produits(id) ON DELETE CASCADE,
            type_mouvement  TEXT    NOT NULL,
            quantite        REAL    NOT NULL DEFAULT 0.0,
            stock_avant     REAL    DEFAULT 0.0,
            stock_apres     REAL    DEFAULT 0.0,
            reference       TEXT    DEFAULT '',
            note            TEXT    DEFAULT '',
            utilisateur_id  INTEGER REFERENCES utilisateurs(id) ON DELETE SET NULL,
            date_mouvement  TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 13. Inventaires ──────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS inventaires (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            utilisateur_id  INTEGER REFERENCES utilisateurs(id) ON DELETE SET NULL,
            statut          TEXT    DEFAULT 'en_cours',
            note            TEXT    DEFAULT '',
            date_inventaire TEXT    DEFAULT (datetime('now')),
            date_cloture    TEXT    DEFAULT NULL
        )
    """)

    # ── 14. Lignes d'inventaire ──────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS lignes_inventaire (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            inventaire_id   INTEGER NOT NULL REFERENCES inventaires(id) ON DELETE CASCADE,
            produit_id      INTEGER NOT NULL REFERENCES produits(id) ON DELETE RESTRICT,
            stock_theorique REAL    DEFAULT 0.0,
            stock_reel      REAL    DEFAULT 0.0,
            ecart           REAL    DEFAULT 0.0
        )
    """)

    # ── 15. Arrêtés de compte ────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS arretes_compte (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            reference           TEXT    DEFAULT '',
            periode_debut       TEXT    DEFAULT '',
            periode_fin         TEXT    DEFAULT '',
            ca_brut             REAL    DEFAULT 0.0,
            nb_ventes           INTEGER DEFAULT 0,
            nb_ventes_annulees  INTEGER DEFAULT 0,
            montant_annule      REAL    DEFAULT 0.0,
            total_depenses      REAL    DEFAULT 0.0,
            benefice_net        REAL    DEFAULT 0.0,
            stock_valeur        REAL    DEFAULT 0.0,
            total_especes       REAL    DEFAULT 0.0,
            total_credit        REAL    DEFAULT 0.0,
            total_mobile        REAL    DEFAULT 0.0,
            operateur           TEXT    DEFAULT '',
            notes               TEXT    DEFAULT '',
            utilisateur_id      INTEGER REFERENCES utilisateurs(id) ON DELETE SET NULL,
            date_creation       TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 16. Dépenses ─────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS depenses (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            libelle         TEXT    NOT NULL,
            montant         REAL    NOT NULL DEFAULT 0.0,
            categorie       TEXT    DEFAULT '',
            utilisateur_id  INTEGER REFERENCES utilisateurs(id) ON DELETE SET NULL,
            note            TEXT    DEFAULT '',
            date_depense    TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 17. Journal (log opérations) ─────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS journal (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            utilisateur_id  INTEGER REFERENCES utilisateurs(id) ON DELETE SET NULL,
            action          TEXT    NOT NULL,
            detail          TEXT    DEFAULT '',
            date_action     TEXT    DEFAULT (datetime('now'))
        )
    """)

    # ── 18. Commandes clients ─────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS commandes_clients (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            reference       TEXT    NOT NULL UNIQUE,
            client_nom      TEXT    NOT NULL,
            client_tel      TEXT    DEFAULT '',
            note            TEXT    DEFAULT '',
            total           REAL    DEFAULT 0.0,
            statut          TEXT    DEFAULT 'en_attente',
            operateur       TEXT    DEFAULT '',
            date_commande   TEXT    DEFAULT (datetime('now')),
            date_livraison  TEXT    DEFAULT NULL
        )
    """)

    # ── 19. Lignes commande client ────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS lignes_commande_client (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            commande_id     INTEGER NOT NULL REFERENCES commandes_clients(id) ON DELETE CASCADE,
            produit_id      INTEGER NOT NULL REFERENCES produits(id) ON DELETE RESTRICT,
            produit_nom     TEXT    NOT NULL,
            quantite        REAL    NOT NULL DEFAULT 1.0,
            prix_unitaire   REAL    NOT NULL DEFAULT 0.0
        )
    """)

    # ── Créer admin par défaut si table vide ────────────────────────────────
    cur.execute("SELECT COUNT(*) FROM utilisateurs")
    if cur.fetchone()[0] == 0:
        import hashlib
        SALT = "HSHOP_SECURE_2026_V1"
        h = hashlib.sha256(("ADMIN" + SALT).encode()).hexdigest()
        cur.execute("""
            INSERT INTO utilisateurs (identifiant, mot_de_passe, role)
            VALUES ('admin', ?, 'admin')
        """, (h,))
        log.info("Compte admin par défaut créé — mdp: ADMIN")

    conn.commit()

    # ── Index de performance ─────────────────────────────────────────────────
    _creer_indexes(cur, conn)

    # ── Migrations ───────────────────────────────────────────────────────────
    _migrations(cur, conn)

    conn.close()
    log.info("Base de données initialisée avec succès.")


# ═══════════════════════════════════════════════════════════════════════════
# INDEX DE PERFORMANCE
# ═══════════════════════════════════════════════════════════════════════════

def _creer_indexes(cur, conn):
    indexes = [
        # Ventes — requêtes les plus fréquentes
        "CREATE INDEX IF NOT EXISTS idx_ventes_date        ON ventes(date_vente)",
        "CREATE INDEX IF NOT EXISTS idx_ventes_client      ON ventes(client_id)",
        "CREATE INDEX IF NOT EXISTS idx_ventes_annulee     ON ventes(annulee)",
        "CREATE INDEX IF NOT EXISTS idx_ventes_cloture     ON ventes(cloture)",
        "CREATE INDEX IF NOT EXISTS idx_ventes_mode        ON ventes(mode_paiement)",
        "CREATE INDEX IF NOT EXISTS idx_ventes_vendeur     ON ventes(vendeur)",
        # Produits
        "CREATE INDEX IF NOT EXISTS idx_produits_nom       ON produits(nom)",
        "CREATE INDEX IF NOT EXISTS idx_produits_categorie ON produits(categorie_id)",
        "CREATE INDEX IF NOT EXISTS idx_produits_actif     ON produits(actif)",
        # Mouvements stock
        "CREATE INDEX IF NOT EXISTS idx_mouvements_prod    ON mouvements_stock(produit_id)",
        "CREATE INDEX IF NOT EXISTS idx_mouvements_date    ON mouvements_stock(date_mouvement)",
        "CREATE INDEX IF NOT EXISTS idx_mouvements_type    ON mouvements_stock(type_mouvement)",
        "CREATE INDEX IF NOT EXISTS idx_mouvements_vente   ON mouvements_stock(vente_id)",
        # Clients
        "CREATE INDEX IF NOT EXISTS idx_clients_nom        ON clients(nom)",
        # Commandes clients
        "CREATE INDEX IF NOT EXISTS idx_cmd_clients_statut ON commandes_clients(statut)",
        "CREATE INDEX IF NOT EXISTS idx_cmd_clients_date   ON commandes_clients(date_commande)",
        "CREATE INDEX IF NOT EXISTS idx_lignes_cmd_client  ON lignes_commande_client(commande_id)",
        # Dépenses
        "CREATE INDEX IF NOT EXISTS idx_depenses_date      ON depenses(date_depense)",
        "CREATE INDEX IF NOT EXISTS idx_depenses_cloture   ON depenses(cloture)",
    ]
    # Index optionnels — colonnes pouvant varier selon version DB
    optional = [
        "CREATE INDEX IF NOT EXISTS idx_echeances_statut   ON echeances_paiement(statut)",
        "CREATE INDEX IF NOT EXISTS idx_echeances_date     ON echeances_paiement(date_echeance)",
        "CREATE INDEX IF NOT EXISTS idx_echeances_vente    ON echeances_paiement(vente_id)",
        "CREATE INDEX IF NOT EXISTS idx_reglements_vente   ON reglements_credit(vente_id)",
        "CREATE INDEX IF NOT EXISTS idx_lignes_cmd_prod    ON lignes_commande(produit_id)",
        "CREATE INDEX IF NOT EXISTS idx_produits_code      ON produits(code_barre)",
        "CREATE INDEX IF NOT EXISTS idx_produits_code2     ON produits(code)",
    ]
    for idx in indexes + optional:
        try:
            cur.execute(idx)
        except Exception as e:
            log.debug(f"Index ignoré : {e}")
    conn.commit()


# ═══════════════════════════════════════════════════════════════════════════
# MIGRATIONS — ajout de colonnes sur bases existantes
# ═══════════════════════════════════════════════════════════════════════════

def _migrations(cur, conn):
    migrations = [
        # echeances_paiement
        "ALTER TABLE echeances_paiement ADD COLUMN statut         TEXT DEFAULT 'en_attente'",
        "ALTER TABLE echeances_paiement ADD COLUMN date_reglement TEXT DEFAULT NULL",
        "ALTER TABLE echeances_paiement ADD COLUMN note           TEXT DEFAULT ''",
        # clients
        "ALTER TABLE clients ADD COLUMN encours_max  REAL DEFAULT 0.0",
        "ALTER TABLE clients ADD COLUMN solde_credit REAL DEFAULT 0.0",
        # produits
        "ALTER TABLE produits ADD COLUMN categorie_id  INTEGER DEFAULT NULL",
        "ALTER TABLE produits ADD COLUMN fournisseur_id INTEGER DEFAULT NULL",
        "ALTER TABLE produits ADD COLUMN stock_min   REAL DEFAULT 0.0",
        "ALTER TABLE produits ADD COLUMN stock_actuel REAL DEFAULT 0.0",
        "ALTER TABLE produits ADD COLUMN seuil_alerte REAL DEFAULT 0.0",
        "ALTER TABLE produits ADD COLUMN code        TEXT DEFAULT ''",
        "ALTER TABLE produits ADD COLUMN unite       TEXT DEFAULT 'pce'",
        "ALTER TABLE produits ADD COLUMN actif       INTEGER DEFAULT 1",
        "ALTER TABLE produits ADD COLUMN updated_at  TEXT DEFAULT (datetime('now'))",
        # ventes — colonnes critiques
        "ALTER TABLE ventes ADD COLUMN annulee          INTEGER DEFAULT 0",
        "ALTER TABLE ventes ADD COLUMN motif_annulation TEXT DEFAULT ''",
        "ALTER TABLE ventes ADD COLUMN montant_regle    REAL DEFAULT 0.0",
        "ALTER TABLE ventes ADD COLUMN remise           REAL DEFAULT 0.0",
        "ALTER TABLE ventes ADD COLUMN note             TEXT DEFAULT ''",
        "ALTER TABLE ventes ADD COLUMN client_id        INTEGER DEFAULT NULL",
        "ALTER TABLE ventes ADD COLUMN cloture          INTEGER DEFAULT 0",
        # mouvements_stock
        "ALTER TABLE mouvements_stock ADD COLUMN vente_id  INTEGER DEFAULT NULL",
        "ALTER TABLE mouvements_stock ADD COLUMN cloture   INTEGER DEFAULT 0",
        # reglements_credit
        "ALTER TABLE reglements_credit ADD COLUMN vente_id INTEGER DEFAULT NULL",
        "ALTER TABLE reglements_credit ADD COLUMN cloture  INTEGER DEFAULT 0",
        # depenses
        "ALTER TABLE depenses ADD COLUMN cloture     INTEGER DEFAULT 0",
        "ALTER TABLE depenses ADD COLUMN categorie   TEXT DEFAULT ''",
        # arretes_compte — colonnes complètes utilisées par tab_arrete.py
        "ALTER TABLE arretes_compte ADD COLUMN total_mobile        REAL    DEFAULT 0.0",
        "ALTER TABLE arretes_compte ADD COLUMN reference           TEXT    DEFAULT ''",
        "ALTER TABLE arretes_compte ADD COLUMN periode_debut       TEXT    DEFAULT ''",
        "ALTER TABLE arretes_compte ADD COLUMN periode_fin         TEXT    DEFAULT ''",
        "ALTER TABLE arretes_compte ADD COLUMN ca_brut             REAL    DEFAULT 0.0",
        "ALTER TABLE arretes_compte ADD COLUMN nb_ventes           INTEGER DEFAULT 0",
        "ALTER TABLE arretes_compte ADD COLUMN nb_ventes_annulees  INTEGER DEFAULT 0",
        "ALTER TABLE arretes_compte ADD COLUMN montant_annule      REAL    DEFAULT 0.0",
        "ALTER TABLE arretes_compte ADD COLUMN total_depenses      REAL    DEFAULT 0.0",
        "ALTER TABLE arretes_compte ADD COLUMN benefice_net        REAL    DEFAULT 0.0",
        "ALTER TABLE arretes_compte ADD COLUMN stock_valeur        REAL    DEFAULT 0.0",
        "ALTER TABLE arretes_compte ADD COLUMN operateur           TEXT    DEFAULT ''",
        "ALTER TABLE arretes_compte ADD COLUMN date_creation       TEXT    DEFAULT (datetime('now'))",
        # arretes_compte — distingue une clôture perso (un caissier, ses
        # propres ventes) d'une clôture globale (admin, toutes les caisses)
        "ALTER TABLE arretes_compte ADD COLUMN perimetre           TEXT    DEFAULT 'globale'",
        # fournisseurs — colonnes utilisées par tab_fournisseurs.py
        "ALTER TABLE fournisseurs ADD COLUMN contact TEXT DEFAULT ''",
        "ALTER TABLE fournisseurs ADD COLUMN notes   TEXT DEFAULT ''",
        # lignes_inventaire — colonnes utilisées par tab_inventaire.py (valider/enregistrer la saisie)
        "ALTER TABLE lignes_inventaire ADD COLUMN ecart        REAL DEFAULT 0.0",
        "ALTER TABLE lignes_inventaire ADD COLUMN valeur_ecart REAL DEFAULT 0.0",
        "ALTER TABLE lignes_inventaire ADD COLUMN notes        TEXT DEFAULT ''",
        # lignes_commande — colonne utilisée par tab_commandes.py (réception partielle)
        "ALTER TABLE lignes_commande ADD COLUMN quantite_recue REAL DEFAULT 0.0",
        # commandes_fournisseurs — colonnes utilisées par tab_commandes.py (réception / règlement)
        "ALTER TABLE commandes_fournisseurs ADD COLUMN date_reception TEXT DEFAULT NULL",
        "ALTER TABLE commandes_fournisseurs ADD COLUMN montant_paye   REAL DEFAULT 0.0",
    ]
    for sql in migrations:
        try:
            cur.execute(sql)
            conn.commit()
        except Exception:
            pass  # Colonne déjà existante — normal


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS MÉTIER
# ═══════════════════════════════════════════════════════════════════════════

def enregistrer_mouvement(conn_or_produit_id, produit_id_or_type=None,
                           type_or_quantite=None, quantite_or_operateur=None,
                           operateur_or_none=None, **kwargs):
    """
    Enregistre un mouvement de stock. Accepte deux signatures :
    1. (conn, produit_id, type_mouvement, quantite, operateur, **kwargs)
    2. (produit_id, type_mouvement, quantite, **kwargs) — kwargs: produit_id=, type_mouvement=, quantite=
    """
    import sqlite3 as _sqlite3

    # Signature kwargs nommés (tab_commandes_clients) ou (None, ...)
    if 'produit_id' in kwargs or conn_or_produit_id is None:
        produit_id    = kwargs.pop('produit_id', produit_id_or_type)
        type_mv       = kwargs.pop('type_mouvement', type_or_quantite or 'ENTREE')
        quantite      = kwargs.pop('quantite', quantite_or_operateur or 0)
        operateur     = kwargs.pop('operateur', operateur_or_none or '')
        conn_externe  = None
        own_conn      = True
    elif isinstance(conn_or_produit_id, _sqlite3.Connection):
        # (conn, produit_id, type, quantite, operateur, **kwargs)
        conn_externe  = conn_or_produit_id
        produit_id    = produit_id_or_type
        type_mv       = type_or_quantite
        quantite      = quantite_or_operateur
        operateur     = operateur_or_none or kwargs.pop('operateur', '')
        own_conn      = False
    else:
        # (produit_id, type, quantite, operateur, **kwargs)
        produit_id    = conn_or_produit_id
        type_mv       = produit_id_or_type
        quantite      = type_or_quantite
        operateur     = quantite_or_operateur or kwargs.pop('operateur', '')
        conn_externe  = None
        own_conn      = True

    prix_unitaire   = float(kwargs.pop('prix_unitaire',   0.0) or 0.0)
    vente_id        = kwargs.pop('vente_id',        None)
    reference_doc   = kwargs.pop('reference_doc',   kwargs.pop('reference', ''))
    motif           = kwargs.pop('motif',           kwargs.pop('note', ''))
    fournisseur_id  = kwargs.pop('fournisseur_id',  None)
    fournisseur_nom = kwargs.pop('fournisseur_nom', '')

    try:
        if own_conn:
            conn = get_connection()
        else:
            conn = conn_externe

        cur = conn.cursor()
        cur.execute("SELECT stock, stock_actuel, nom, code_barre FROM produits WHERE id=?",
                    (produit_id,))
        row = cur.fetchone()
        if not row:
            if own_conn:
                conn.close()
            return

        # stock_actuel est la colonne de référence partout ailleurs (web,
        # rapports) — on la préfère désormais à "stock" en cas de divergence
        # entre les deux (bug corrigé le 20/06/2026 : l'ordre inverse causait
        # des mouvements calculés sur la mauvaise valeur de départ).
        stock_avant  = float(row[1] or row[0] or 0)
        produit_nom  = row[2] or ""
        produit_code = row[3] or ""
        qte          = float(quantite or 0)
        qte_abs      = abs(qte)

        sorties = ("SORTIE_VENTE", "SORTIE_MANUELLE", "sortie", "vente",
                   "inventaire_moins", "perte")
        if type_mv in sorties or qte < 0:
            stock_apres = stock_avant - qte_abs
            qte_insert  = -qte_abs
        else:
            stock_apres = stock_avant + qte_abs
            qte_insert  = qte_abs

        montant = qte_abs * prix_unitaire

        cur.execute("""
            INSERT INTO mouvements_stock
                (produit_id, produit_code, produit_nom, type_mouvement,
                 quantite, stock_avant, stock_apres,
                 prix_unitaire, montant_total,
                 fournisseur_id, fournisseur_nom,
                 vente_id, reference_doc, operateur, motif)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (produit_id, produit_code, produit_nom, type_mv,
              qte_insert, stock_avant, stock_apres,
              prix_unitaire, montant,
              fournisseur_id, fournisseur_nom,
              vente_id, reference_doc, operateur, motif))

        # Mettre à jour stock (updated_at optionnel)
        try:
            cur.execute(
                "UPDATE produits SET stock=?, stock_actuel=?, updated_at=datetime('now') WHERE id=?",
                (stock_apres, stock_apres, produit_id))
        except Exception:
            cur.execute(
                "UPDATE produits SET stock=?, stock_actuel=? WHERE id=?",
                (stock_apres, stock_apres, produit_id))

        if own_conn:
            conn.commit()
            conn.close()

    except Exception as e:
        log.error(f"enregistrer_mouvement: {e}")
        if own_conn:
            try:
                conn.rollback()
                conn.close()
            except Exception:
                pass


def enregistrer_journal(utilisateur_id: int, action: str, detail: str = ""):
    """Ajoute une entrée dans le journal des opérations."""
    try:
        conn = get_connection()
        conn.execute("""
            INSERT INTO journal (utilisateur_id, action, detail)
            VALUES (?, ?, ?)
        """, (utilisateur_id, action, detail))
        conn.commit()
        conn.close()
    except Exception as e:
        log.error(f"enregistrer_journal: {e}")


def get_stock_critique() -> list:
    """Retourne les produits dont stock <= stock_min."""
    conn = get_connection()
    cur  = conn.cursor()
    cur.execute("""
        SELECT p.id, p.nom, p.stock, p.stock_min, c.nom AS categorie
        FROM produits p
        LEFT JOIN categories c ON c.id = p.categorie_id
        WHERE p.stock <= p.stock_min AND p.actif = 1
        ORDER BY (p.stock - p.stock_min) ASC
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


# ── Alias de compatibilité ───────────────────────────────────────────────────
initialiser_db = init_db   # ancien nom utilisé dans main.py

def migrate_db():
    """
    Applique les migrations sur une DB existante.
    À appeler à chaque démarrage pour garantir la compatibilité.
    """
    conn = get_connection()
    cur  = conn.cursor()
    _migrations(cur, conn)
    conn.close()
    log.info("Migrations DB appliquées.")



# ── Fonctions héritées — compatibilité ui_caisse.py ─────────────────────────

def couleur_stock(stock: float, stock_min: float = 0.0) -> tuple:
    """Retourne (tag, couleur_hex) selon le niveau de stock."""
    if stock <= 0:
        return ("rupture",  "#e8394d")   # rouge — rupture
    if stock_min > 0 and stock <= stock_min:
        return ("critique", "#f5a623")   # orange — stock critique
    return ("ok", "#00c896")             # vert — stock OK


def appliquer_tags_stock(tree, stock_col: str = "stock",
                          stock_min_col: str = "stock_min"):
    """
    Applique des tags de couleur sur les lignes d'un Treeview
    selon le niveau de stock. Appelle tree.tag_configure() si besoin.
    """
    try:
        tree.tag_configure("rupture",  foreground="#e8394d")
        tree.tag_configure("critique", foreground="#f5a623")
        tree.tag_configure("ok",       foreground="#00c896")
        for iid in tree.get_children():
            vals = tree.item(iid, "values")
            cols = list(tree["columns"])
            if stock_col not in cols:
                continue
            idx_s  = cols.index(stock_col)
            idx_sm = cols.index(stock_min_col) if stock_min_col in cols else -1
            try:
                stock    = float(str(vals[idx_s]).replace(",", "").replace(" ", "") or 0)
                stock_min = float(str(vals[idx_sm]).replace(",", "").replace(" ", "") or 0) if idx_sm >= 0 else 0
            except (ValueError, IndexError):
                continue
            if stock <= 0:
                tag = "rupture"
            elif stock_min > 0 and stock <= stock_min:
                tag = "critique"
            else:
                tag = "ok"
            tree.item(iid, tags=(tag,))
    except Exception as e:
        log.warning(f"appliquer_tags_stock: {e}")


def exporter_stock_excel(chemin: str = None) -> str:
    """Exporte le stock complet vers un fichier Excel. Retourne le chemin."""
    try:
        import openpyxl
        from datetime import datetime
        if not chemin:
            dossier = os.path.join(os.path.dirname(get_db_path()), "exports")
            os.makedirs(dossier, exist_ok=True)
            chemin = os.path.join(dossier, f"stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.code_barre, p.nom, c.nom AS categorie,
                   p.stock, p.stock_min, p.unite,
                   p.prix_achat, p.prix_vente
            FROM produits p
            LEFT JOIN categories c ON c.id = p.categorie_id
            WHERE p.actif = 1
            ORDER BY p.nom
        """)
        rows = cur.fetchall()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock"
        headers = ["Code-barres", "Désignation", "Catégorie", "Stock",
                   "Stock min", "Unité", "Prix achat", "Prix vente"]
        ws.append(headers)
        for row in rows:
            ws.append(list(row))
        wb.save(chemin)
        log.info(f"Stock exporté : {chemin}")
        return chemin
    except Exception as e:
        log.error(f"exporter_stock_excel: {e}")
        return ""


def importer_stock_excel(chemin: str) -> tuple:
    """
    Importe les produits depuis un fichier Excel.
    Retourne (nb_importes, nb_erreurs, messages).
    Colonnes attendues : code_barre, nom, categorie, stock, stock_min, unite,
                         prix_achat, prix_vente
    """
    nb_ok, nb_err, msgs = 0, 0, []
    try:
        import openpyxl
        wb   = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        conn = get_connection()
        cur  = conn.cursor()
        for i, row in enumerate(rows, 2):
            try:
                if not row or not row[1]:
                    continue
                code_barre, nom, categorie, stock, stock_min, unite, prix_achat, prix_vente = (
                    (row[j] if j < len(row) else None) for j in range(8)
                )
                # Catégorie
                cat_id = None
                if categorie:
                    cur.execute("SELECT id FROM categories WHERE nom=?", (str(categorie),))
                    r = cur.fetchone()
                    if r:
                        cat_id = r["id"]
                    else:
                        cur.execute("INSERT INTO categories(nom) VALUES(?)", (str(categorie),))
                        cat_id = cur.lastrowid

                # Vérifier si l'article existe déjà (par code_barre ou nom)
                existing_id = None
                if code_barre:
                    cur.execute("SELECT id FROM produits WHERE code_barre=?", (str(code_barre),))
                    r = cur.fetchone()
                    if r: existing_id = r["id"]
                if not existing_id:
                    cur.execute("SELECT id FROM produits WHERE nom=?", (str(nom),))
                    r = cur.fetchone()
                    if r: existing_id = r["id"]

                if existing_id:
                    cur.execute("""
                        UPDATE produits SET
                            nom=?, categorie_id=?, stock=?, stock_min=?,
                            unite=?, prix_achat=?, prix_vente=?,
                            updated_at=datetime('now')
                        WHERE id=?
                    """, (str(nom), cat_id, float(stock or 0), float(stock_min or 0),
                             str(unite or "pce"), float(prix_achat or 0), float(prix_vente or 0),
                             existing_id))
                else:
                    cur.execute("""
                        INSERT INTO produits
                            (code_barre, nom, categorie_id, stock, stock_min, unite, prix_achat, prix_vente)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                    str(code_barre or ""), str(nom), cat_id,
                    float(stock or 0), float(stock_min or 0),
                    str(unite or "pce"),
                    float(prix_achat or 0), float(prix_vente or 0)
                ))
                nb_ok += 1
            except Exception as e:
                nb_err += 1
                msgs.append(f"Ligne {i} : {e}")
        conn.commit()
        conn.close()
        log.info(f"Import Excel : {nb_ok} OK, {nb_err} erreurs")
    except Exception as e:
        log.error(f"importer_stock_excel: {e}")
        msgs.append(str(e))
    return nb_ok, nb_err, msgs


def get_impaye_total(client_id: int) -> float:
    """Retourne le montant total impayé d'un client."""
    conn = get_connection()
    cur  = conn.cursor()
    cur.execute("""
        SELECT COALESCE(SUM(e.montant), 0)
        FROM echeances_paiement e
        JOIN ventes v ON v.id = e.vente_id
        WHERE v.client_id = ?
          AND e.statut IN ('en_attente', 'partiel')
          AND e.date_echeance < DATE('now')
    """, (client_id,))
    total = cur.fetchone()[0] or 0.0
    conn.close()
    return total

def lire_produits():
    """Récupère tous les produits de la base de données."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        # Ajustez les noms des colonnes selon votre vraie table
        cursor.execute("SELECT id, nom, prix, quantite FROM produits")
        lignes = cursor.fetchall()
        
        # On transforme le résultat en liste de dictionnaires
        produits = []
        for ligne in lignes:
            produits.append({
                "id": ligne[0],
                "nom": ligne[1],
                "prix": ligne[2],
                "quantite": ligne[3]
            })
        return produits
    except sqlite3.Error as e:
        log.error(f"Erreur SQL : {e}")
        return None
    finally:
        conn.close()

def ajouter_produit(nom, prix, quantite):
    """Insère un nouveau produit dans la base."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO produits (nom, prix, quantite) VALUES (?, ?, ?)",
            (nom, prix, quantite)
        )
        conn.commit()
        return True
    except sqlite3.Error as e:
        log.error(f"Erreur d'insertion : {e}")
        return False
    finally:
        conn.close()
