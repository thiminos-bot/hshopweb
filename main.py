# -*- coding: utf-8 -*-
"""
main.py — HSHOP V2.1 WEB API
BIOS CREATIONS — Bobo-Dioulasso

Cette API ne réinvente pas la logique métier : elle s'appuie sur
database.py (connexion, synchro stock/stock_actuel, journal des
mouvements) et fec.py (facturation électronique certifiée DGI) —
exactement les mêmes modules que la version desktop. C'est ce qui
garantit que les deux versions restent cohérentes sur la même base.
"""
import base64
import os
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from contextlib import asynccontextmanager
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timedelta
import jwt
import uvicorn
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

import config
import database
import fec
import impression

# Initialise / met à jour le schéma au démarrage (idempotent — IF NOT EXISTS
# + migrations ALTER TABLE déjà gérées par database.py). Remplace l'ancien
# @app.on_event("startup"), déprécié depuis les versions récentes de FastAPI.
@asynccontextmanager
async def lifespan(app: FastAPI):
    database.init_db()
    database.migrate_db()
    config.logger.info("API HSHOP démarrée — base initialisée/vérifiée.")
    yield

app = FastAPI(title="HSHOP_API_WEB", version="2.1", lifespan=lifespan)

# Dossier contenant ce fichier — c'est aussi là que doivent se trouver
# caisse.html, login.html, historique.html, fec.py, database.py, config.py.
WEB_DIR = os.path.dirname(os.path.abspath(__file__))

# --- PAGES HTML (servies par l'API elle-même : même origine que /produits,
# /ventes, etc. → plus besoin de CORS ni d'URL en dur dans le JS) ---
@app.get("/")
def serve_root():
    return FileResponse(os.path.join(WEB_DIR, "login.html"))

@app.get("/login.html")
def serve_login():
    return FileResponse(os.path.join(WEB_DIR, "login.html"))

@app.get("/caisse.html")
def serve_caisse():
    return FileResponse(os.path.join(WEB_DIR, "caisse.html"))

@app.get("/historique.html")
def serve_historique():
    return FileResponse(os.path.join(WEB_DIR, "historique.html"))

@app.get("/dashboard.html")
def serve_dashboard():
    return FileResponse(os.path.join(WEB_DIR, "dashboard.html"))

@app.get("/produits.html")
def serve_produits_page():
    return FileResponse(os.path.join(WEB_DIR, "produits.html"))

@app.get("/clients.html")
def serve_clients_page():
    return FileResponse(os.path.join(WEB_DIR, "clients.html"))

@app.get("/parametres.html")
def serve_parametres_page():
    return FileResponse(os.path.join(WEB_DIR, "parametres.html"))

@app.get("/inventaire.html")
def serve_inventaire_page():
    return FileResponse(os.path.join(WEB_DIR, "inventaire.html"))

@app.get("/cloture-caisse.html")
def serve_cloture_caisse_page():
    return FileResponse(os.path.join(WEB_DIR, "cloture-caisse.html"))

@app.get("/mouvements.html")
def serve_mouvements_page():
    return FileResponse(os.path.join(WEB_DIR, "mouvements.html"))

@app.get("/journal.html")
def serve_journal_page():
    return FileResponse(os.path.join(WEB_DIR, "journal.html"))

@app.get("/arrete-mensuel.html")
def serve_arrete_mensuel_page():
    return FileResponse(os.path.join(WEB_DIR, "arrete-mensuel.html"))

@app.get("/utilisateurs.html")
def serve_utilisateurs_page():
    return FileResponse(os.path.join(WEB_DIR, "utilisateurs.html"))

@app.get("/commandes.html")
def serve_commandes_page():
    return FileResponse(os.path.join(WEB_DIR, "commandes.html"))

@app.get("/logo.png")
def serve_logo():
    chemin = config.get_logo_path()
    if not chemin:
        raise HTTPException(status_code=404, detail="Logo non configuré.")
    return FileResponse(chemin, media_type="image/png")

# Configuration du CORS
# Plus de wildcard "*" — les pages sont désormais servies par cette même API,
# donc les appels fetch() sont en same-origin et ne passent même pas par le
# CORS du navigateur. Cette liste ne sert que si un jour un autre frontend
# (sur une autre origine) doit appeler cette API directement.
ALLOWED_ORIGINS: list[str] = [
    # ex: "http://100.112.59.66:8080" si un autre outil doit appeler l'API
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SECRET_KEY = config.SALT
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# --- SCHEMAS PYDANTIC ---
class Token(BaseModel):
    access_token: str
    token_type: str
    role: str

class LigneVenteSchema(BaseModel):
    produit_id: int
    quantite: float
    prix_unitaire: float
    remise: float = 0.0

class VenteCreateSchema(BaseModel):
    client_id: Optional[int] = None
    mode_paiement: str = "ESPECES"
    remise: float = 0.0
    recu: float = 0.0
    note: Optional[str] = ""
    lignes: List[LigneVenteSchema]

class ClientSchema(BaseModel):
    nom: str
    telephone: str = ""
    email: str = ""
    adresse: str = ""
    encours_max: float = 0.0

class ReglementSchema(BaseModel):
    montant: float
    mode_paiement: str = "ESPECES"
    note: str = ""

class ParametresSchema(BaseModel):
    nom_structure: str
    adresse_structure: str = ""
    tel_structure: str = ""
    ifu_structure: str = ""
    rccm_structure: str = ""

class HorairesCaisseSchema(BaseModel):
    actif: bool
    heure_ouverture: str = "08:00"
    heure_fermeture: str = "18:00"

class ImprimanteConfigSchema(BaseModel):
    active: bool
    methode: str = "win"  # "win" (driver Windows/USB), "lpt", "usb_serie"
    nom_imprimante: str = ""
    port_serie: str = "COM1"
    baudrate: int = 9600
    largeur: int = 32

ROLES_VALIDES = ("admin", "vendeur", "magasinier")

class UtilisateurCreateSchema(BaseModel):
    identifiant: str
    mot_de_passe: str
    role: str = "vendeur"

class UtilisateurUpdateSchema(BaseModel):
    role: str
    actif: bool

class ResetMotDePasseSchema(BaseModel):
    nouveau_mot_de_passe: str

class LigneCommandeSchema(BaseModel):
    produit_id: int
    quantite: float
    prix_unitaire: float

class CommandeCreateSchema(BaseModel):
    fournisseur_id: int
    note: str = ""
    date_livraison: Optional[str] = None
    lignes: List[LigneCommandeSchema]

class ReceptionLigneSchema(BaseModel):
    ligne_id: int
    quantite_recue: float  # quantité reçue LORS DE CETTE réception (pas le cumul)

class ReceptionCommandeSchema(BaseModel):
    lignes: List[ReceptionLigneSchema]

class PaiementCommandeSchema(BaseModel):
    montant: float

class DemarrerInventaireSchema(BaseModel):
    note: str = ""

class ClotureCaisseSchema(BaseModel):
    notes: str = ""
    perimetre: str = "perso"  # "perso" (ses propres ventes) ou "globale" (admin uniquement)

class DepenseSchema(BaseModel):
    categorie: str = ""
    description: str = ""
    montant: float
    mode_paiement: str = "ESPECES"
    beneficiaire: str = ""
    notes: str = ""

class LigneComptageSchema(BaseModel):
    stock_reel: Optional[float] = None  # None = remettre la ligne en "non compté"
    notes: str = ""

class ApprovisionnerSchema(BaseModel):
    quantite: float

class ProduitCreateSchema(BaseModel):
    nom: str
    code_barre: str = ""
    code: str = ""
    prix_achat: float = 0.0
    prix_vente: float
    stock_actuel: float = 0.0
    stock_min: float = 0.0
    seuil_alerte: float = 0.0
    unite: str = "pce"
    description: str = ""
    categorie_id: Optional[int] = None
    fournisseur_id: Optional[int] = None

class ProduitUpdateSchema(BaseModel):
    nom: str
    code_barre: str = ""
    code: str = ""
    prix_achat: float = 0.0
    prix_vente: float
    stock_min: float = 0.0
    seuil_alerte: float = 0.0
    unite: str = "pce"
    description: str = ""
    categorie_id: Optional[int] = None
    fournisseur_id: Optional[int] = None

class ProduitStatutSchema(BaseModel):
    actif: bool

class CategorieSchema(BaseModel):
    nom: str
    description: str = ""

class FournisseurSchema(BaseModel):
    nom: str
    contact: str = ""
    telephone: str = ""
    email: str = ""
    adresse: str = ""
    notes: str = ""

# --- DEPENDANCE DE SECURITE STRICTE ---
# Important : contrairement à l'ancienne version, l'absence de token ou un
# token invalide renvoie désormais une vraie 401 — il n'y a plus de
# "bypass_local" qui laisse entrer n'importe qui sans authentification.
def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Session invalide ou expirée. Merci de vous reconnecter.",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: Optional[str] = payload.get("sub")
        role: Optional[str] = payload.get("role")
        if username is None:
            raise credentials_exception
        return {"username": username, "role": role}
    except jwt.PyJWTError:
        raise credentials_exception

# Le dashboard manager expose des données sensibles (CA, crédits clients) —
# réservé au rôle admin, pas vendeur/magasinier.
def get_current_admin(current_user: dict = Depends(get_current_user)) -> dict:
    if current_user.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Accès réservé aux comptes administrateur.",
        )
    return current_user

def _journaliser(username: str, action: str, detail: str = ""):
    """Trace une action dans le journal d'audit. N'échoue jamais bruyamment :
    un souci de journalisation ne doit jamais faire échouer l'action elle-même."""
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id FROM utilisateurs WHERE identifiant = ?", (username,))
        u = cur.fetchone()
        conn.close()
        database.enregistrer_journal(u["id"] if u else None, action, detail)
    except Exception as e:
        config.logger.warning(f"Journalisation échouée pour '{action}': {e}")

# --- ROUTE 1 : AUTHENTIFICATION ---
@app.post("/token", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends()):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT mot_de_passe, role, actif FROM utilisateurs WHERE identifiant = ?", (form_data.username,))
        user = cur.fetchone()
        conn.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not user or user["actif"] == 0 or not config.verifier_mdp(form_data.password, user["mot_de_passe"]):
        raise HTTPException(status_code=401, detail="Identifiant ou mot de passe incorrect.")

    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode = {"sub": form_data.username, "role": user["role"], "exp": expire}
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    _journaliser(form_data.username, "CONNEXION", "Connexion réussie depuis l'application web")
    return {"access_token": encoded_jwt, "token_type": "bearer", "role": user["role"]}

# --- ROUTE 2 : LISTE DES PRODUITS ---
# NOTE : ne filtre plus "stock_actuel > 0" — c'était ce qui faisait
# disparaître les produits en rupture de l'onglet Stock (impossible de les
# réapprovisionner). Le filtre de stock se fait maintenant côté frontend,
# uniquement pour la grille de vente.
@app.get("/produits")
def lister_tous_les_produits(current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, nom, prix_vente, stock_actuel, unite, code_barre FROM produits WHERE actif = 1")
        produits = cur.fetchall()
        resultat = [{"id": p["id"], "nom": p["nom"], "prix_vente": p["prix_vente"], "stock": p["stock_actuel"],
                     "unite": p["unite"] if p["unite"] else "pce", "code_barre": p["code_barre"] or ""} for p in produits]
        conn.close()
        return resultat
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- ROUTE 2bis : CRÉER UN PRODUIT (utilisée par l'import CSV) ---
@app.post("/produits", status_code=status.HTTP_201_CREATED)
def creer_produit(data: ProduitCreateSchema, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO produits (nom, code_barre, code, prix_achat, prix_vente, stock, stock_actuel,
                                   stock_min, seuil_alerte, unite, description, categorie_id, fournisseur_id, actif)
            VALUES (?, ?, ?, ?, ?, 0, 0, ?, ?, ?, ?, ?, ?, 1)
        """, (data.nom, data.code_barre, data.code, data.prix_achat, data.prix_vente,
              data.stock_min, data.seuil_alerte, data.unite, data.description,
              data.categorie_id, data.fournisseur_id))
        produit_id = cur.lastrowid
        conn.commit()

        # Stock initial enregistré comme un vrai mouvement (traçable dans
        # mouvements_stock), pas comme une valeur insérée en silence.
        if data.stock_actuel > 0:
            database.enregistrer_mouvement(
                conn, produit_id, "ENTREE_INITIALE", data.stock_actuel,
                current_user["username"], reference_doc="Création produit Web"
            )
            conn.commit()
        conn.close()
        return {"status": "success", "produit_id": produit_id}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- GESTION PRODUITS : liste complète (actifs + inactifs) pour l'admin ---
@app.get("/produits/gestion")
def lister_produits_gestion(current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT p.id, p.nom, p.code_barre, p.code, p.prix_achat, p.prix_vente,
                   p.stock_actuel, p.stock_min, p.seuil_alerte, p.unite, p.description,
                   p.actif, p.categorie_id, p.fournisseur_id,
                   c.nom as categorie_nom, f.nom as fournisseur_nom
            FROM produits p
            LEFT JOIN categories c ON c.id = p.categorie_id
            LEFT JOIN fournisseurs f ON f.id = p.fournisseur_id
            ORDER BY p.nom ASC
        """)
        rows = cur.fetchall()
        conn.close()
        return [dict(r) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- MODIFIER UN PRODUIT (jamais le stock — passe par /approvisionner) ---
@app.put("/produits/{produit_id}")
def modifier_produit(produit_id: int, data: ProduitUpdateSchema, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id FROM produits WHERE id = ?", (produit_id,))
        if not cur.fetchone():
            conn.close()
            raise HTTPException(status_code=404, detail="Produit introuvable.")
        cur.execute("""
            UPDATE produits SET nom=?, code_barre=?, code=?, prix_achat=?, prix_vente=?,
                   stock_min=?, seuil_alerte=?, unite=?, description=?, categorie_id=?, fournisseur_id=?,
                   updated_at=datetime('now','localtime')
            WHERE id=?
        """, (data.nom, data.code_barre, data.code, data.prix_achat, data.prix_vente,
              data.stock_min, data.seuil_alerte, data.unite, data.description,
              data.categorie_id, data.fournisseur_id, produit_id))
        conn.commit()
        conn.close()
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- ACTIVER / DESACTIVER UN PRODUIT (suppression douce) ---
@app.patch("/produits/{produit_id}/statut")
def changer_statut_produit(produit_id: int, data: ProduitStatutSchema, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, nom FROM produits WHERE id = ?", (produit_id,))
        produit = cur.fetchone()
        if not produit:
            conn.close()
            raise HTTPException(status_code=404, detail="Produit introuvable.")
        cur.execute("UPDATE produits SET actif=?, updated_at=datetime('now','localtime') WHERE id=?",
                    (1 if data.actif else 0, produit_id))
        conn.commit()
        conn.close()
        _journaliser(current_user["username"], "PRODUIT_ACTIF" if data.actif else "PRODUIT_INACTIF", produit["nom"])
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- CATEGORIES ---
@app.get("/categories")
def lister_categories(current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, nom, description FROM categories ORDER BY nom ASC")
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/categories", status_code=status.HTTP_201_CREATED)
def creer_categorie(data: CategorieSchema, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO categories (nom, description) VALUES (?, ?)", (data.nom, data.description))
        conn.commit()
        cat_id = cur.lastrowid
        conn.close()
        return {"status": "success", "id": cat_id}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/categories/{categorie_id}")
def modifier_categorie(categorie_id: int, data: CategorieSchema, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("UPDATE categories SET nom=?, description=? WHERE id=?", (data.nom, data.description, categorie_id))
        if cur.rowcount == 0:
            conn.close()
            raise HTTPException(status_code=404, detail="Catégorie introuvable.")
        conn.commit()
        conn.close()
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/categories/{categorie_id}")
def supprimer_categorie(categorie_id: int, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        # ON DELETE SET NULL sur produits.categorie_id : les produits liés
        # ne sont pas supprimés, juste détachés de la catégorie.
        cur.execute("DELETE FROM categories WHERE id=?", (categorie_id,))
        if cur.rowcount == 0:
            conn.close()
            raise HTTPException(status_code=404, detail="Catégorie introuvable.")
        conn.commit()
        conn.close()
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- FOURNISSEURS ---
@app.get("/fournisseurs")
def lister_fournisseurs(current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, nom, contact, telephone, email, adresse, notes FROM fournisseurs ORDER BY nom ASC")
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/fournisseurs", status_code=status.HTTP_201_CREATED)
def creer_fournisseur(data: FournisseurSchema, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""INSERT INTO fournisseurs (nom, contact, telephone, email, adresse, notes)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (data.nom, data.contact, data.telephone, data.email, data.adresse, data.notes))
        conn.commit()
        f_id = cur.lastrowid
        conn.close()
        return {"status": "success", "id": f_id}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/fournisseurs/{fournisseur_id}")
def modifier_fournisseur(fournisseur_id: int, data: FournisseurSchema, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""UPDATE fournisseurs SET nom=?, contact=?, telephone=?, email=?, adresse=?, notes=?
                       WHERE id=?""",
                    (data.nom, data.contact, data.telephone, data.email, data.adresse, data.notes, fournisseur_id))
        if cur.rowcount == 0:
            conn.close()
            raise HTTPException(status_code=404, detail="Fournisseur introuvable.")
        conn.commit()
        conn.close()
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/fournisseurs/{fournisseur_id}")
def supprimer_fournisseur(fournisseur_id: int, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM fournisseurs WHERE id=?", (fournisseur_id,))
        if cur.rowcount == 0:
            conn.close()
            raise HTTPException(status_code=404, detail="Fournisseur introuvable.")
        conn.commit()
        conn.close()
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- ROUTE 3 : CREER UNE VENTE (+ ÉMISSION FEC) ---
# --- PARAMETRES / IDENTITE DE L'ENTREPRISE ---
# GET ouvert à tout utilisateur connecté : ces infos sont imprimées sur
# chaque ticket, pas confidentielles. Seule la modification est admin-only.
# Route publique (pas de Depends) — uniquement le nom et la présence du logo,
# pour que login.html affiche l'identité avant authentification. Rien de
# sensible (même info que sur un ticket de caisse vu par n'importe quel client).
@app.get("/identite-publique")
def identite_publique():
    cfg = config.charger_app_config()
    return {
        "nom_structure": cfg.get("nom_structure", config.NOM_STRUCTURE),
        "logo_present": config.get_logo_path() is not None,
    }

@app.get("/parametres")
def get_parametres(current_user: dict = Depends(get_current_user)):
    cfg = config.charger_app_config()
    return {
        "nom_structure": cfg.get("nom_structure", config.NOM_STRUCTURE),
        "adresse_structure": cfg.get("adresse_structure", config.ADRESSE_STRUCTURE),
        "tel_structure": cfg.get("tel_structure", config.TEL_STRUCTURE),
        "ifu_structure": cfg.get("ifu_structure", ""),
        "rccm_structure": cfg.get("rccm_structure", ""),
        "logo_present": config.get_logo_path() is not None,
    }

@app.put("/parametres")
def update_parametres(data: ParametresSchema, current_user: dict = Depends(get_current_admin)):
    try:
        cfg = config.charger_app_config()
        cfg["nom_structure"] = data.nom_structure
        cfg["adresse_structure"] = data.adresse_structure
        cfg["tel_structure"] = data.tel_structure
        cfg["ifu_structure"] = data.ifu_structure
        cfg["rccm_structure"] = data.rccm_structure
        config.sauvegarder_app_config(cfg)
        _journaliser(current_user["username"], "PARAMETRES_MODIFIES", f"Identité entreprise mise à jour ({data.nom_structure})")
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/parametres/logo")
async def upload_logo(file: UploadFile = File(...), current_user: dict = Depends(get_current_admin)):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Le fichier doit être une image.")
    contenu = await file.read()
    if len(contenu) > 3 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image trop lourde (3 Mo max).")
    try:
        with open(config.LOGO_PATH, "wb") as f:
            f.write(contenu)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- HORAIRES DE CAISSE / FERMETURE AUTOMATIQUE ---
# Lecture ouverte à tous les connectés : le JS de chaque page (caisse, stock,
# clients...) doit pouvoir lire l'heure de fermeture pour avertir/déconnecter
# automatiquement les caissiers. Seule la modification est admin-only.
@app.get("/horaires-caisse")
def get_horaires_caisse(current_user: dict = Depends(get_current_user)):
    return config.charger_horaires_caisse()

@app.put("/horaires-caisse")
def update_horaires_caisse(data: HorairesCaisseSchema, current_user: dict = Depends(get_current_admin)):
    try:
        config.sauver_horaires_caisse({
            "actif": data.actif,
            "heure_ouverture": data.heure_ouverture,
            "heure_fermeture": data.heure_fermeture,
        })
        _journaliser(current_user["username"], "HORAIRES_MODIFIES",
                     f"actif={data.actif}, fermeture={data.heure_fermeture}")
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/fermeture-auto.js")
def serve_fermeture_auto_js():
    return FileResponse(os.path.join(WEB_DIR, "fermeture-auto.js"), media_type="application/javascript")

# --- IMPRIMANTE THERMIQUE (impression directe ESC/POS, admin uniquement) ---
METHODES_IMPRIMANTE_VALIDES = ("win", "lpt", "usb_serie")

@app.get("/imprimante/config")
def get_config_imprimante(current_user: dict = Depends(get_current_admin)):
    return impression.charger_config_imprimante()

@app.put("/imprimante/config")
def update_config_imprimante(data: ImprimanteConfigSchema, current_user: dict = Depends(get_current_admin)):
    if data.methode not in METHODES_IMPRIMANTE_VALIDES:
        raise HTTPException(status_code=400, detail=f"Méthode invalide — choisir parmi : {', '.join(METHODES_IMPRIMANTE_VALIDES)}.")
    try:
        impression.sauver_config_imprimante({
            "active": data.active, "methode": data.methode, "nom_imprimante": data.nom_imprimante,
            "port_serie": data.port_serie, "baudrate": data.baudrate, "largeur": data.largeur,
        })
        _journaliser(current_user["username"], "IMPRIMANTE_CONFIGUREE", f"active={data.active}, méthode={data.methode}")
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/imprimante/diagnostic")
def diagnostic_imprimante(current_user: dict = Depends(get_current_admin)):
    return {"diagnostic": impression.diagnostiquer_impression()}

@app.post("/imprimante/test")
def test_imprimante(current_user: dict = Depends(get_current_admin)):
    cfg = impression.charger_config_imprimante()
    if not cfg.get("active"):
        raise HTTPException(status_code=400, detail="Active d'abord l'impression et enregistre la configuration avant de tester.")
    try:
        impression.imprimer_ticket_thermique(
            vendeur=current_user["username"],
            panier=[{"nom": "Article de test", "qty": 1, "prix": 1000, "remise": 0}],
            total=1000, recu=1000, monnaie=0, id_vente=0, mode="TEST", client_nom=None,
        )
        return {"status": "success", "message": "Commande d'impression envoyée — vérifie que le ticket sort bien. En cas d'échec silencieux, regarde hshop_v2.log."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- INVENTAIRE ---
def _ligne_inventaire_to_dict(r):
    return {
        "id": r["id"], "produit_id": r["produit_id"], "produit_nom": r["nom"], "unite": r["unite"],
        "stock_theorique": r["stock_theorique"], "stock_reel": r["stock_reel"],
        "ecart": r["ecart"], "valeur_ecart": r["valeur_ecart"], "notes": r["notes"],
    }

@app.get("/inventaire/courant")
def inventaire_courant(current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, note, date_inventaire FROM inventaires WHERE statut='EN_COURS' ORDER BY id DESC LIMIT 1")
        inv = cur.fetchone()
        if not inv:
            conn.close()
            return {"actif": False}
        cur.execute("""
            SELECT li.id, li.produit_id, li.stock_theorique, li.stock_reel, li.ecart, li.valeur_ecart, li.notes,
                   p.nom, p.unite
            FROM lignes_inventaire li JOIN produits p ON p.id = li.produit_id
            WHERE li.inventaire_id = ? ORDER BY p.nom ASC
        """, (inv["id"],))
        lignes = [_ligne_inventaire_to_dict(r) for r in cur.fetchall()]
        conn.close()
        return {"actif": True, "inventaire": {"id": inv["id"], "note": inv["note"], "date_inventaire": inv["date_inventaire"]}, "lignes": lignes}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/inventaire/demarrer", status_code=status.HTTP_201_CREATED)
def demarrer_inventaire(data: DemarrerInventaireSchema, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id FROM inventaires WHERE statut='EN_COURS' LIMIT 1")
        existant = cur.fetchone()
        if existant:
            conn.close()
            raise HTTPException(status_code=400, detail=f"Un inventaire est déjà en cours (#{existant['id']}) — reprends-le au lieu d'en créer un nouveau.")

        cur.execute("SELECT id FROM utilisateurs WHERE identifiant = ?", (current_user["username"],))
        u = cur.fetchone()
        utilisateur_id = u["id"] if u else None

        cur.execute("INSERT INTO inventaires (utilisateur_id, statut, note) VALUES (?, 'EN_COURS', ?)", (utilisateur_id, data.note))
        inventaire_id = cur.lastrowid

        cur.execute("SELECT id, stock_actuel FROM produits WHERE actif = 1")
        produits = cur.fetchall()
        for p in produits:
            cur.execute("""
                INSERT INTO lignes_inventaire (inventaire_id, produit_id, stock_theorique, stock_reel, ecart, valeur_ecart, notes)
                VALUES (?, ?, ?, NULL, 0, 0, '')
            """, (inventaire_id, p["id"], p["stock_actuel"]))

        conn.commit()
        conn.close()
        return {"status": "success", "inventaire_id": inventaire_id, "nb_produits": len(produits)}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/inventaire/lignes/{ligne_id}")
def compter_ligne_inventaire(ligne_id: int, data: LigneComptageSchema, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT li.id, li.stock_theorique, li.inventaire_id, p.prix_achat, i.statut
            FROM lignes_inventaire li
            JOIN produits p ON p.id = li.produit_id
            JOIN inventaires i ON i.id = li.inventaire_id
            WHERE li.id = ?
        """, (ligne_id,))
        ligne = cur.fetchone()
        if not ligne:
            conn.close()
            raise HTTPException(status_code=404, detail="Ligne d'inventaire introuvable.")
        if ligne["statut"] != "EN_COURS":
            conn.close()
            raise HTTPException(status_code=400, detail="Cet inventaire est déjà clôturé, le comptage n'est plus modifiable.")

        if data.stock_reel is None:
            ecart, valeur_ecart, stock_reel = 0.0, 0.0, None
        else:
            ecart = data.stock_reel - ligne["stock_theorique"]
            valeur_ecart = ecart * (ligne["prix_achat"] or 0)
            stock_reel = data.stock_reel

        cur.execute("UPDATE lignes_inventaire SET stock_reel=?, ecart=?, valeur_ecart=?, notes=? WHERE id=?",
                    (stock_reel, ecart, valeur_ecart, data.notes, ligne_id))
        conn.commit()
        conn.close()
        return {"status": "success", "ecart": ecart, "valeur_ecart": valeur_ecart}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/inventaire/{inventaire_id}/cloturer")
def cloturer_inventaire(inventaire_id: int, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, statut FROM inventaires WHERE id = ?", (inventaire_id,))
        inv = cur.fetchone()
        if not inv:
            conn.close()
            raise HTTPException(status_code=404, detail="Inventaire introuvable.")
        if inv["statut"] != "EN_COURS":
            conn.close()
            raise HTTPException(status_code=400, detail="Cet inventaire est déjà clôturé.")

        cur.execute("SELECT id, produit_id, stock_reel FROM lignes_inventaire WHERE inventaire_id = ? AND stock_reel IS NOT NULL", (inventaire_id,))
        lignes = cur.fetchall()

        nb_ajustements = 0
        for ligne in lignes:
            cur.execute("SELECT stock_actuel FROM produits WHERE id = ?", (ligne["produit_id"],))
            stock_actuel_courant = cur.fetchone()["stock_actuel"]
            # On compare au stock ACTUEL (pas au théorique capturé au démarrage) :
            # ça absorbe les ventes/mouvements survenus pendant le comptage et
            # garantit que le stock final = exactement ce qui a été compté.
            delta = ligne["stock_reel"] - stock_actuel_courant
            if delta < 0:
                database.enregistrer_mouvement(conn, ligne["produit_id"], "inventaire_moins", abs(delta),
                                                current_user["username"], reference_doc=f"INVENTAIRE-{inventaire_id}")
                nb_ajustements += 1
            elif delta > 0:
                database.enregistrer_mouvement(conn, ligne["produit_id"], "inventaire_plus", delta,
                                                current_user["username"], reference_doc=f"INVENTAIRE-{inventaire_id}")
                nb_ajustements += 1

        cur.execute("UPDATE inventaires SET statut='CLOTURE', date_cloture=datetime('now','localtime') WHERE id=?", (inventaire_id,))
        conn.commit()
        conn.close()
        return {"status": "success", "nb_ajustements": nb_ajustements}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/inventaire/historique")
def historique_inventaires(current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, note, date_inventaire, date_cloture FROM inventaires WHERE statut='CLOTURE' ORDER BY date_cloture DESC LIMIT 30")
        invs = cur.fetchall()
        resultat = []
        for inv in invs:
            cur.execute("""
                SELECT SUM(CASE WHEN ecart != 0 THEN 1 ELSE 0 END) as nb_ecarts, COALESCE(SUM(valeur_ecart),0) as total
                FROM lignes_inventaire WHERE inventaire_id = ?
            """, (inv["id"],))
            agg = cur.fetchone()
            resultat.append({
                "id": inv["id"], "note": inv["note"], "date_inventaire": inv["date_inventaire"], "date_cloture": inv["date_cloture"],
                "nb_ecarts": agg["nb_ecarts"] or 0, "valeur_totale_ecart": agg["total"],
            })
        conn.close()
        return resultat
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/inventaire/{inventaire_id}")
def detail_inventaire(inventaire_id: int, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, statut, note, date_inventaire, date_cloture FROM inventaires WHERE id = ?", (inventaire_id,))
        inv = cur.fetchone()
        if not inv:
            conn.close()
            raise HTTPException(status_code=404, detail="Inventaire introuvable.")
        cur.execute("""
            SELECT li.id, li.produit_id, li.stock_theorique, li.stock_reel, li.ecart, li.valeur_ecart, li.notes,
                   p.nom, p.unite
            FROM lignes_inventaire li JOIN produits p ON p.id = li.produit_id
            WHERE li.inventaire_id = ? ORDER BY p.nom ASC
        """, (inventaire_id,))
        lignes = [_ligne_inventaire_to_dict(r) for r in cur.fetchall()]
        conn.close()
        return {"inventaire": dict(inv), "lignes": lignes}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/inventaire/{inventaire_id}/export-excel")
def exporter_inventaire_excel(inventaire_id: int, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, note, date_inventaire, date_cloture FROM inventaires WHERE id = ?", (inventaire_id,))
        inv = cur.fetchone()
        if not inv:
            conn.close()
            raise HTTPException(status_code=404, detail="Inventaire introuvable.")
        cur.execute("""
            SELECT p.nom, p.unite, li.stock_theorique, li.stock_reel, li.ecart, li.valeur_ecart, li.notes
            FROM lignes_inventaire li JOIN produits p ON p.id = li.produit_id
            WHERE li.inventaire_id = ? ORDER BY p.nom ASC
        """, (inventaire_id,))
        lignes = cur.fetchall()
        conn.close()

        cfg = config.charger_app_config()
        nom_structure = cfg.get("nom_structure", config.NOM_STRUCTURE)

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventaire"

        ws.merge_cells("A1:G1")
        ws["A1"] = nom_structure
        ws["A1"].font = Font(bold=True, size=14)

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Inventaire #{inv['id']}" + (f" — {inv['note']}" if inv["note"] else "")
        ws["A2"].font = Font(italic=True, size=10, color="666666")

        ws.merge_cells("A3:G3")
        statut_txt = f"Démarré le {inv['date_inventaire']}"
        statut_txt += f" — Clôturé le {inv['date_cloture']}" if inv["date_cloture"] else " — EN COURS"
        ws["A3"] = statut_txt
        ws["A3"].font = Font(size=10, color="666666")

        headers = ["Produit", "Unité", "Stock théorique", "Stock compté", "Écart", "Valeur écart (F)", "Notes"]
        header_row = 5
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="1F2937")
            cell.alignment = Alignment(horizontal="center")

        row = header_row + 1
        total_valeur_ecart = 0.0
        nb_ecarts = 0
        for l in lignes:
            compte = l["stock_reel"] is not None
            ws.cell(row=row, column=1, value=l["nom"])
            ws.cell(row=row, column=2, value=l["unite"])
            ws.cell(row=row, column=3, value=l["stock_theorique"])
            ws.cell(row=row, column=4, value=l["stock_reel"] if compte else "—")
            ecart_cell = ws.cell(row=row, column=5, value=l["ecart"] if compte else "—")
            valeur_cell = ws.cell(row=row, column=6, value=l["valeur_ecart"] if compte else "—")
            ws.cell(row=row, column=7, value=l["notes"])
            if compte and l["ecart"] != 0:
                couleur = "DC2626" if l["ecart"] < 0 else "16A34A"
                ecart_cell.font = Font(color=couleur, bold=True)
                valeur_cell.font = Font(color=couleur, bold=True)
                nb_ecarts += 1
                total_valeur_ecart += l["valeur_ecart"]
            ws.cell(row=row, column=3).number_format = "#,##0.##"
            if compte:
                ws.cell(row=row, column=4).number_format = "#,##0.##"
                ecart_cell.number_format = "+#,##0.##;-#,##0.##"
                valeur_cell.number_format = '#,##0 "F";-#,##0 "F"'
            row += 1

        row += 1
        ws.cell(row=row, column=4, value="TOTAL ÉCARTS :").font = Font(bold=True)
        ws.cell(row=row, column=5, value=nb_ecarts).font = Font(bold=True)
        total_cell = ws.cell(row=row, column=6, value=total_valeur_ecart)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0 "F";-#,##0 "F"'

        for i, w in enumerate([28, 8, 14, 14, 10, 16, 25], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = f"A{header_row + 1}"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="inventaire_{inventaire_id}.xlsx"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- DEPENSES ---
@app.get("/depenses")
def lister_depenses(current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        if current_user.get("role") == "admin":
            cur.execute("""
                SELECT id, date_depense, categorie, description, montant, mode_paiement, beneficiaire, operateur, notes
                FROM depenses WHERE cloture=0 ORDER BY date_depense DESC
            """)
        else:
            cur.execute("""
                SELECT id, date_depense, categorie, description, montant, mode_paiement, beneficiaire, operateur, notes
                FROM depenses WHERE cloture=0 AND operateur=? ORDER BY date_depense DESC
            """, (current_user["username"],))
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/depenses/historique")
def historique_depenses(current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        if current_user.get("role") == "admin":
            cur.execute("""
                SELECT id, date_depense, categorie, description, montant, mode_paiement, beneficiaire, operateur, notes, cloture
                FROM depenses ORDER BY date_depense DESC LIMIT 200
            """)
        else:
            cur.execute("""
                SELECT id, date_depense, categorie, description, montant, mode_paiement, beneficiaire, operateur, notes, cloture
                FROM depenses WHERE operateur=? ORDER BY date_depense DESC LIMIT 200
            """, (current_user["username"],))
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/depenses", status_code=status.HTTP_201_CREATED)
def creer_depense(data: DepenseSchema, current_user: dict = Depends(get_current_user)):
    if data.montant <= 0:
        raise HTTPException(status_code=400, detail="Le montant doit être positif.")
    if not data.description.strip():
        raise HTTPException(status_code=400, detail="Le motif (description) est obligatoire.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO depenses (libelle, categorie, description, montant, mode_paiement, beneficiaire, operateur, notes, cloture)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0)
        """, (data.description, data.categorie, data.description, data.montant, data.mode_paiement, data.beneficiaire, current_user["username"], data.notes))
        conn.commit()
        depense_id = cur.lastrowid
        conn.close()
        return {"status": "success", "id": depense_id}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/depenses/{depense_id}")
def supprimer_depense(depense_id: int, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT operateur, cloture, description, montant FROM depenses WHERE id=?", (depense_id,))
        d = cur.fetchone()
        if not d:
            conn.close()
            raise HTTPException(status_code=404, detail="Dépense introuvable.")
        if d["cloture"]:
            conn.close()
            raise HTTPException(status_code=400, detail="Cette dépense a déjà été incluse dans une clôture, elle ne peut plus être supprimée.")
        if current_user.get("role") != "admin" and d["operateur"] != current_user["username"]:
            conn.close()
            raise HTTPException(status_code=403, detail="Tu ne peux supprimer que tes propres dépenses.")
        cur.execute("DELETE FROM depenses WHERE id=?", (depense_id,))
        conn.commit()
        conn.close()
        _journaliser(current_user["username"], "DEPENSE_SUPPRIMEE", f"{d['description']} — {d['montant']:.0f} F")
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- CLOTURE DE CAISSE ---
def _calculer_periode_ouverte(cur, vendeur_filtre: Optional[str] = None):
    """Calcule les totaux de la période non encore clôturée. Si vendeur_filtre
    est fourni, ne porte que sur les ventes/dépenses de ce caissier."""
    params_ventes = []
    filtre_v = ""
    if vendeur_filtre:
        filtre_v = "AND vendeur = ?"
        params_ventes = [vendeur_filtre]

    cur.execute(f"SELECT COUNT(*), COALESCE(SUM(total),0) FROM ventes WHERE cloture=0 AND annulee=0 {filtre_v}", params_ventes)
    nb_ventes, ca_brut = cur.fetchone()
    cur.execute(f"SELECT COUNT(*), COALESCE(SUM(total),0) FROM ventes WHERE cloture=0 AND annulee=1 {filtre_v}", params_ventes)
    nb_ventes_annulees, montant_annule = cur.fetchone()
    cur.execute(f"SELECT mode_paiement, COALESCE(SUM(total),0) FROM ventes WHERE cloture=0 AND annulee=0 {filtre_v} GROUP BY mode_paiement", params_ventes)
    par_mode = {r[0]: r[1] for r in cur.fetchall()}
    total_especes = par_mode.get("ESPECES", 0.0)
    total_credit = par_mode.get("CREDIT", 0.0)
    total_mobile = par_mode.get("MOBILE", 0.0)

    if vendeur_filtre:
        cur.execute("SELECT COALESCE(SUM(montant),0) FROM depenses WHERE cloture=0 AND operateur = ?", (vendeur_filtre,))
    else:
        cur.execute("SELECT COALESCE(SUM(montant),0) FROM depenses WHERE cloture=0")
    total_depenses = cur.fetchone()[0]

    cur.execute(f"""
        SELECT COALESCE(SUM((lv.prix_unitaire - p.prix_achat) * lv.quantite),0)
        FROM lignes_vente lv JOIN ventes v ON v.id = lv.vente_id JOIN produits p ON p.id = lv.produit_id
        WHERE v.cloture=0 AND v.annulee=0 {filtre_v.replace('vendeur', 'v.vendeur') if filtre_v else ''}
    """, params_ventes)
    marge_brute = cur.fetchone()[0]
    benefice_net = marge_brute - total_depenses
    cur.execute("SELECT COALESCE(SUM(stock_actuel * prix_achat),0) FROM produits WHERE actif=1")
    stock_valeur = cur.fetchone()[0]
    cur.execute(f"SELECT MIN(date_vente) FROM ventes WHERE cloture=0 {filtre_v}", params_ventes)
    periode_debut = cur.fetchone()[0]

    return {
        "nb_ventes": nb_ventes, "ca_brut": ca_brut,
        "nb_ventes_annulees": nb_ventes_annulees, "montant_annule": montant_annule,
        "total_especes": total_especes, "total_credit": total_credit, "total_mobile": total_mobile,
        "total_depenses": total_depenses, "benefice_net": benefice_net, "stock_valeur": stock_valeur,
        "net_caisse": total_especes + total_mobile - total_depenses,
        "periode_debut": periode_debut,
    }

@app.get("/cloture-caisse/courant")
def cloture_caisse_courant(perimetre: str = "perso", current_user: dict = Depends(get_current_user)):
    if perimetre == "globale" and current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Seul un administrateur peut voir la caisse globale.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        vendeur_filtre = None if perimetre == "globale" else current_user["username"]
        totaux = _calculer_periode_ouverte(cur, vendeur_filtre)
        cur.execute("SELECT reference, date_creation FROM arretes_compte ORDER BY id DESC LIMIT 1")
        derniere = cur.fetchone()
        conn.close()
        totaux["perimetre"] = perimetre
        totaux["derniere_cloture"] = dict(derniere) if derniere else None
        return totaux
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/cloture-caisse", status_code=status.HTTP_201_CREATED)
def cloturer_caisse(data: ClotureCaisseSchema, current_user: dict = Depends(get_current_user)):
    if data.perimetre == "globale" and current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Seul un administrateur peut clôturer la caisse globale.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        vendeur_filtre = None if data.perimetre == "globale" else current_user["username"]
        totaux = _calculer_periode_ouverte(cur, vendeur_filtre)

        if totaux["nb_ventes"] == 0 and totaux["total_depenses"] == 0:
            conn.close()
            raise HTTPException(status_code=400, detail="Rien à clôturer pour ce périmètre.")

        cur.execute("SELECT id FROM utilisateurs WHERE identifiant = ?", (current_user["username"],))
        u = cur.fetchone()
        utilisateur_id = u["id"] if u else None

        maintenant = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        suffixe = current_user["username"] if vendeur_filtre else "GLOBAL"
        reference = f"ARRETE-{datetime.now():%Y%m%d-%H%M%S}-{suffixe}"
        periode_debut = totaux["periode_debut"] or maintenant

        cur.execute("""
            INSERT INTO arretes_compte
                (reference, periode_debut, periode_fin, ca_brut, nb_ventes, nb_ventes_annulees, montant_annule,
                 total_depenses, benefice_net, stock_valeur, total_especes, total_credit, total_mobile,
                 operateur, notes, utilisateur_id, perimetre)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (reference, periode_debut, maintenant, totaux["ca_brut"], totaux["nb_ventes"], totaux["nb_ventes_annulees"],
              totaux["montant_annule"], totaux["total_depenses"], totaux["benefice_net"], totaux["stock_valeur"],
              totaux["total_especes"], totaux["total_credit"], totaux["total_mobile"],
              current_user["username"], data.notes, utilisateur_id, data.perimetre))
        arrete_id = cur.lastrowid

        if vendeur_filtre:
            cur.execute("UPDATE ventes SET cloture=1 WHERE cloture=0 AND vendeur=?", (vendeur_filtre,))
            cur.execute("UPDATE depenses SET cloture=1 WHERE cloture=0 AND operateur=?", (vendeur_filtre,))
        else:
            cur.execute("UPDATE ventes SET cloture=1 WHERE cloture=0")
            cur.execute("UPDATE depenses SET cloture=1 WHERE cloture=0")

        conn.commit()
        conn.close()
        _journaliser(current_user["username"], "CLOTURE_CAISSE", f"{reference} — {data.perimetre} — Net caisse : {totaux['net_caisse']:.0f} F")
        return {"status": "success", "id": arrete_id, "reference": reference, **totaux}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/cloture-caisse/historique")
def historique_clotures(current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        if current_user.get("role") == "admin":
            cur.execute("""
                SELECT id, reference, periode_debut, periode_fin, ca_brut, nb_ventes, nb_ventes_annulees,
                       montant_annule, total_depenses, benefice_net, stock_valeur,
                       total_especes, total_credit, total_mobile, operateur, notes, date_creation, perimetre
                FROM arretes_compte ORDER BY id DESC LIMIT 50
            """)
        else:
            # Un caissier ne voit que ses propres clôtures perso, pas celles des autres ni les globales.
            cur.execute("""
                SELECT id, reference, periode_debut, periode_fin, ca_brut, nb_ventes, nb_ventes_annulees,
                       montant_annule, total_depenses, benefice_net, stock_valeur,
                       total_especes, total_credit, total_mobile, operateur, notes, date_creation, perimetre
                FROM arretes_compte WHERE operateur = ? AND perimetre = 'perso' ORDER BY id DESC LIMIT 50
            """, (current_user["username"],))
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- ARRETE MENSUEL (rapport en lecture seule, filtrable — distinct de la clôture) ---
@app.get("/arrete-mensuel/options")
def arrete_mensuel_options(current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT strftime('%Y', date_vente) AS a FROM ventes ORDER BY a DESC")
        annees = [r["a"] for r in cur.fetchall() if r["a"]]
        cur.execute("SELECT DISTINCT vendeur FROM ventes WHERE vendeur IS NOT NULL AND vendeur != '' ORDER BY vendeur")
        vendeurs = [r["vendeur"] for r in cur.fetchall()]
        conn.close()
        return {"annees": annees, "vendeurs": vendeurs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def _calculer_arrete_mensuel(cur, annee: int, mois: int, vendeur: Optional[str] = None, mode_paiement: Optional[str] = None):
    mois_str = f"{annee:04d}-{mois:02d}"

    conditions = ["strftime('%Y-%m', date_vente) = ?"]
    params = [mois_str]
    if vendeur:
        conditions.append("vendeur = ?")
        params.append(vendeur)
    if mode_paiement:
        conditions.append("mode_paiement = ?")
        params.append(mode_paiement)
    base_where = " AND ".join(conditions)

    cur.execute(f"SELECT COUNT(*), COALESCE(SUM(total),0) FROM ventes WHERE {base_where} AND annulee=0", params)
    nb_ventes, ca_brut = cur.fetchone()
    cur.execute(f"SELECT COUNT(*), COALESCE(SUM(total),0) FROM ventes WHERE {base_where} AND annulee=1", params)
    nb_annulees, montant_annule = cur.fetchone()

    cur.execute(f"SELECT mode_paiement, COALESCE(SUM(total),0) FROM ventes WHERE {base_where} AND annulee=0 GROUP BY mode_paiement", params)
    par_mode = {r[0]: r[1] for r in cur.fetchall()}

    # Dépenses du mois (même filtre vendeur -> operateur, même filtre mode de paiement)
    dep_conditions = ["strftime('%Y-%m', date_depense) = ?"]
    dep_params = [mois_str]
    if vendeur:
        dep_conditions.append("operateur = ?")
        dep_params.append(vendeur)
    if mode_paiement:
        dep_conditions.append("mode_paiement = ?")
        dep_params.append(mode_paiement)
    dep_where = " AND ".join(dep_conditions)
    cur.execute(f"SELECT COALESCE(SUM(montant),0) FROM depenses WHERE {dep_where}", dep_params)
    total_depenses = cur.fetchone()[0]
    cur.execute(f"""
        SELECT categorie, COALESCE(SUM(montant),0) FROM depenses WHERE {dep_where}
        GROUP BY categorie ORDER BY 2 DESC
    """, dep_params)
    depenses_par_categorie = [{"categorie": r[0] or "Sans catégorie", "total": r[1]} for r in cur.fetchall()]

    cur.execute(f"""
        SELECT COALESCE(SUM((lv.prix_unitaire - p.prix_achat) * lv.quantite),0)
        FROM lignes_vente lv JOIN ventes v ON v.id = lv.vente_id JOIN produits p ON p.id = lv.produit_id
        WHERE {base_where.replace('date_vente', 'v.date_vente').replace('vendeur', 'v.vendeur').replace('mode_paiement', 'v.mode_paiement')} AND v.annulee=0
    """, params)
    marge_brute = cur.fetchone()[0]
    benefice_net = marge_brute - total_depenses

    # CA par jour du mois (pour le mini-graphique)
    cur.execute(f"""
        SELECT date(date_vente) j, COALESCE(SUM(total),0) FROM ventes WHERE {base_where} AND annulee=0
        GROUP BY j ORDER BY j ASC
    """, params)
    ca_par_jour = [{"jour": r[0], "ca": r[1]} for r in cur.fetchall()]

    # Ventilation par caissier (seulement si pas déjà filtré sur un caissier précis)
    ventilation_vendeurs = []
    if not vendeur:
        conditions_vp = ["strftime('%Y-%m', date_vente) = ?", "annulee=0"]
        params_vp = [mois_str]
        if mode_paiement:
            conditions_vp.append("mode_paiement = ?")
            params_vp.append(mode_paiement)
        cur.execute(f"""
            SELECT vendeur, COUNT(*), COALESCE(SUM(total),0) FROM ventes
            WHERE {" AND ".join(conditions_vp)} GROUP BY vendeur ORDER BY 3 DESC
        """, params_vp)
        ventilation_vendeurs = [{"vendeur": r[0] or "Inconnu", "nb_ventes": r[1], "ca": r[2]} for r in cur.fetchall()]

    # Top 5 produits du mois (respecte les mêmes filtres)
    cur.execute(f"""
        SELECT p.nom, SUM(lv.quantite), SUM(lv.quantite*lv.prix_unitaire)
        FROM lignes_vente lv JOIN ventes v ON v.id = lv.vente_id JOIN produits p ON p.id = lv.produit_id
        WHERE {base_where.replace('date_vente', 'v.date_vente').replace('vendeur', 'v.vendeur').replace('mode_paiement', 'v.mode_paiement')} AND v.annulee=0
        GROUP BY p.id ORDER BY 3 DESC LIMIT 5
    """, params)
    top_produits = [{"nom": r[0], "quantite": r[1], "total": r[2]} for r in cur.fetchall()]

    return {
        "mois": mois_str, "nb_ventes": nb_ventes, "ca_brut": ca_brut,
        "nb_ventes_annulees": nb_annulees, "montant_annule": montant_annule,
        "total_especes": par_mode.get("ESPECES", 0.0), "total_credit": par_mode.get("CREDIT", 0.0), "total_mobile": par_mode.get("MOBILE", 0.0),
        "total_depenses": total_depenses, "depenses_par_categorie": depenses_par_categorie,
        "benefice_net": benefice_net, "ca_par_jour": ca_par_jour,
        "ventilation_vendeurs": ventilation_vendeurs, "top_produits": top_produits,
    }

@app.get("/arrete-mensuel")
def arrete_mensuel(
    annee: int, mois: int, vendeur: Optional[str] = None, mode_paiement: Optional[str] = None,
    current_user: dict = Depends(get_current_admin)
):
    if not (1 <= mois <= 12):
        raise HTTPException(status_code=400, detail="Mois invalide.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        data = _calculer_arrete_mensuel(cur, annee, mois, vendeur, mode_paiement)
        conn.close()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

_NOMS_MOIS = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

@app.get("/arrete-mensuel/export-excel")
def arrete_mensuel_export_excel(
    annee: int, mois: int, vendeur: Optional[str] = None, mode_paiement: Optional[str] = None,
    current_user: dict = Depends(get_current_admin)
):
    if not (1 <= mois <= 12):
        raise HTTPException(status_code=400, detail="Mois invalide.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        d = _calculer_arrete_mensuel(cur, annee, mois, vendeur, mode_paiement)
        conn.close()

        cfg = config.charger_app_config()
        nom_structure = cfg.get("nom_structure", config.NOM_STRUCTURE)
        net_caisse = d["total_especes"] + d["total_mobile"] - d["total_depenses"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Arrêté mensuel"

        ws.merge_cells("A1:C1")
        ws["A1"] = nom_structure
        ws["A1"].font = Font(bold=True, size=14)

        sous_titre = f"Arrêté mensuel — {_NOMS_MOIS[mois]} {annee}"
        filtres = []
        if vendeur: filtres.append(f"Caissier : {vendeur}")
        if mode_paiement: filtres.append(f"Mode : {mode_paiement}")
        if filtres: sous_titre += " — " + " / ".join(filtres)
        ws.merge_cells("A2:C2")
        ws["A2"] = sous_titre
        ws["A2"].font = Font(italic=True, size=10, color="666666")

        def section(titre, ligne):
            cell = ws.cell(row=ligne, column=1, value=titre)
            cell.font = Font(bold=True, color="FFFFFF")
            ws.cell(row=ligne, column=1).fill = PatternFill("solid", start_color="1F2937")
            ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=3)
            return ligne + 1

        def ligne_kv(label, valeur, ligne, gras=False):
            ws.cell(row=ligne, column=1, value=label)
            c = ws.cell(row=ligne, column=2, value=valeur)
            c.number_format = '#,##0 "F";-#,##0 "F"'
            if gras:
                ws.cell(row=ligne, column=1).font = Font(bold=True)
                c.font = Font(bold=True)
            return ligne + 1

        row = 4
        row = section("Activité", row)
        ws.cell(row=row, column=1, value="Nombre de ventes"); ws.cell(row=row, column=2, value=d["nb_ventes"]); row += 1
        row = ligne_kv("Chiffre d'affaires brut", d["ca_brut"], row)
        ws.cell(row=row, column=1, value="Ventes annulées"); ws.cell(row=row, column=2, value=d["nb_ventes_annulees"]); ws.cell(row=row, column=3, value=d["montant_annule"]).number_format = '#,##0 "F"'; row += 1

        row += 1
        row = section("Répartition des encaissements", row)
        row = ligne_kv("Espèces", d["total_especes"], row)
        row = ligne_kv("Mobile Money", d["total_mobile"], row)
        row = ligne_kv("Crédit (non encaissé)", d["total_credit"], row)
        row = ligne_kv("Dépenses", -d["total_depenses"], row)

        row += 1
        row = section("Résultat", row)
        row = ligne_kv("NET CAISSE", net_caisse, row, gras=True)
        row = ligne_kv("Bénéfice net estimé", d["benefice_net"], row, gras=True)

        if d["ventilation_vendeurs"]:
            row += 1
            row = section("Par caissier", row)
            ws.cell(row=row, column=1, value="Caissier"); ws.cell(row=row, column=2, value="Nb ventes"); ws.cell(row=row, column=3, value="CA")
            for c in range(1, 4): ws.cell(row=row, column=c).font = Font(bold=True)
            row += 1
            for v in d["ventilation_vendeurs"]:
                ws.cell(row=row, column=1, value=v["vendeur"])
                ws.cell(row=row, column=2, value=v["nb_ventes"])
                ws.cell(row=row, column=3, value=v["ca"]).number_format = '#,##0 "F"'
                row += 1

        if d["top_produits"]:
            row += 1
            row = section("Top produits", row)
            for p in d["top_produits"]:
                ws.cell(row=row, column=1, value=p["nom"])
                ws.cell(row=row, column=2, value=p["quantite"])
                ws.cell(row=row, column=3, value=p["total"]).number_format = '#,##0 "F"'
                row += 1

        if d["depenses_par_categorie"]:
            row += 1
            row = section("Dépenses par catégorie", row)
            for c in d["depenses_par_categorie"]:
                ws.cell(row=row, column=1, value=c["categorie"])
                ws.cell(row=row, column=3, value=c["total"]).number_format = '#,##0 "F"'
                row += 1

        for i, w in enumerate([30, 16, 16], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="arrete_mensuel_{annee}_{mois:02d}.xlsx"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/arrete-mensuel/export-pdf")
def arrete_mensuel_export_pdf(
    annee: int, mois: int, vendeur: Optional[str] = None, mode_paiement: Optional[str] = None,
    current_user: dict = Depends(get_current_admin)
):
    if not (1 <= mois <= 12):
        raise HTTPException(status_code=400, detail="Mois invalide.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        d = _calculer_arrete_mensuel(cur, annee, mois, vendeur, mode_paiement)
        conn.close()

        cfg = config.charger_app_config()
        nom_structure = cfg.get("nom_structure", config.NOM_STRUCTURE)
        adresse_structure = cfg.get("adresse_structure", config.ADRESSE_STRUCTURE)
        net_caisse = d["total_especes"] + d["total_mobile"] - d["total_depenses"]

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()
        titre_style = ParagraphStyle("TitreStructure", parent=styles["Title"], fontSize=16, spaceAfter=2)
        sous_titre_style = ParagraphStyle("SousTitre", parent=styles["Normal"], textColor=colors.grey, spaceAfter=14)
        section_style = ParagraphStyle("Section", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)

        filtres = []
        if vendeur: filtres.append(f"Caissier : {vendeur}")
        if mode_paiement: filtres.append(f"Mode de paiement : {mode_paiement}")
        ligne_filtres = " — ".join(filtres) if filtres else "Tous caissiers, tous modes de paiement"

        story = [
            Paragraph(nom_structure, titre_style),
            Paragraph(adresse_structure, sous_titre_style),
            Paragraph(f"ARRÊTÉ MENSUEL — {_NOMS_MOIS[mois]} {annee}", styles["Heading1"]),
            Paragraph(ligne_filtres, styles["Normal"]),
            Spacer(1, 10),
        ]

        def table_simple(data_rows, gras=False):
            t = Table(data_rows, colWidths=[100 * mm, 60 * mm])
            style = [
                ("FONTSIZE", (0, 0), (-1, -1), 11 if gras else 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8 if gras else 6),
                ("TOPPADDING", (0, 0), (-1, -1), 8 if gras else 2),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ]
            if gras:
                style += [("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"), ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F3F4F6"))]
            else:
                style.append(("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#E5E7EB")))
            t.setStyle(TableStyle(style))
            return t

        story.append(Paragraph("Activité", section_style))
        story.append(table_simple([
            ["Nombre de ventes", str(d["nb_ventes"])],
            ["Chiffre d'affaires brut", f"{d['ca_brut']:,.0f} F".replace(",", " ")],
            ["Ventes annulées", f"{d['nb_ventes_annulees']} ({d['montant_annule']:,.0f} F)".replace(",", " ")],
        ]))

        story.append(Paragraph("Répartition des encaissements", section_style))
        story.append(table_simple([
            ["Espèces", f"{d['total_especes']:,.0f} F".replace(",", " ")],
            ["Mobile Money", f"{d['total_mobile']:,.0f} F".replace(",", " ")],
            ["Crédit (non encaissé)", f"{d['total_credit']:,.0f} F".replace(",", " ")],
            ["Dépenses", f"-{d['total_depenses']:,.0f} F".replace(",", " ")],
        ]))

        story.append(Spacer(1, 8))
        story.append(table_simple([
            ["NET CAISSE (espèces + mobile − dépenses)", f"{net_caisse:,.0f} F".replace(",", " ")],
            ["Bénéfice net estimé", f"{d['benefice_net']:,.0f} F".replace(",", " ")],
        ], gras=True))

        if d["ventilation_vendeurs"]:
            story.append(Paragraph("Par caissier", section_style))
            rows = [["Caissier", "Nb ventes", "CA"]] + [[v["vendeur"], str(v["nb_ventes"]), f"{v['ca']:,.0f} F".replace(",", " ")] for v in d["ventilation_vendeurs"]]
            t = Table(rows, colWidths=[70 * mm, 40 * mm, 50 * mm])
            t.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 10), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ]))
            story.append(t)

        if d["top_produits"]:
            story.append(Paragraph("Top produits", section_style))
            rows = [["Produit", "Qté", "Total"]] + [[p["nom"], str(p["quantite"]), f"{p['total']:,.0f} F".replace(",", " ")] for p in d["top_produits"]]
            t = Table(rows, colWidths=[70 * mm, 40 * mm, 50 * mm])
            t.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 10), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ]))
            story.append(t)

        story.append(Spacer(1, 20))
        story.append(Paragraph(f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — HSHOP Web", sous_titre_style))

        doc.build(story)
        buffer.seek(0)
        return StreamingResponse(
            buffer, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="arrete_mensuel_{annee}_{mois:02d}.pdf"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/cloture-caisse/{arrete_id}/pdf")
def exporter_cloture_pdf(arrete_id: int, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM arretes_compte WHERE id = ?", (arrete_id,))
        a = cur.fetchone()
        conn.close()
        if not a:
            raise HTTPException(status_code=404, detail="Clôture introuvable.")
        if current_user.get("role") != "admin" and not (a["perimetre"] == "perso" and a["operateur"] == current_user["username"]):
            raise HTTPException(status_code=403, detail="Tu ne peux télécharger que tes propres clôtures.")

        cfg = config.charger_app_config()
        nom_structure = cfg.get("nom_structure", config.NOM_STRUCTURE)
        adresse_structure = cfg.get("adresse_structure", config.ADRESSE_STRUCTURE)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()
        titre_style = ParagraphStyle("TitreStructure", parent=styles["Title"], fontSize=16, spaceAfter=2)
        sous_titre_style = ParagraphStyle("SousTitre", parent=styles["Normal"], textColor=colors.grey, spaceAfter=14)
        section_style = ParagraphStyle("Section", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)

        story = [
            Paragraph(nom_structure, titre_style),
            Paragraph(adresse_structure, sous_titre_style),
            Paragraph(f"ARRÊTÉ DE CAISSE — {a['reference']}", styles["Heading1"]),
            Paragraph(f"Périmètre : {'Caisse personnelle — ' + a['operateur'] if a['perimetre'] == 'perso' else 'Toutes les caisses (consolidé)'}", styles["Normal"]),
            Paragraph(f"Période : du {a['periode_debut']} au {a['periode_fin']}", styles["Normal"]),
            Paragraph(f"Clôturé par : {a['operateur']}", styles["Normal"]),
            Spacer(1, 10),
        ]

        story.append(Paragraph("Activité", section_style))
        data_activite = [
            ["Nombre de ventes", str(a["nb_ventes"])],
            ["Chiffre d'affaires brut", f"{a['ca_brut']:,.0f} F".replace(",", " ")],
            ["Ventes annulées", f"{a['nb_ventes_annulees']} ({a['montant_annule']:,.0f} F)".replace(",", " ")],
        ]
        t1 = Table(data_activite, colWidths=[100 * mm, 60 * mm])
        t1.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#E5E7EB")),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ]))
        story.append(t1)

        story.append(Paragraph("Répartition des encaissements", section_style))
        data_modes = [
            ["Espèces", f"{a['total_especes']:,.0f} F".replace(",", " ")],
            ["Mobile Money", f"{a['total_mobile']:,.0f} F".replace(",", " ")],
            ["Crédit (non encaissé)", f"{a['total_credit']:,.0f} F".replace(",", " ")],
            ["Dépenses", f"-{a['total_depenses']:,.0f} F".replace(",", " ")],
        ]
        t2 = Table(data_modes, colWidths=[100 * mm, 60 * mm])
        t2.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#E5E7EB")),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ]))
        story.append(t2)

        net_caisse = a["total_especes"] + a["total_mobile"] - a["total_depenses"]
        story.append(Spacer(1, 8))
        data_resultat = [
            ["NET CAISSE (espèces + mobile − dépenses)", f"{net_caisse:,.0f} F".replace(",", " ")],
            ["Bénéfice net estimé", f"{a['benefice_net']:,.0f} F".replace(",", " ")],
            ["Valeur du stock (au prix d'achat)", f"{a['stock_valeur']:,.0f} F".replace(",", " ")],
        ]
        t3 = Table(data_resultat, colWidths=[100 * mm, 60 * mm])
        t3.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F3F4F6")),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ]))
        story.append(t3)

        if a["notes"]:
            story.append(Paragraph("Notes", section_style))
            story.append(Paragraph(a["notes"], styles["Normal"]))

        story.append(Spacer(1, 20))
        story.append(Paragraph(f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — HSHOP Web", sous_titre_style))

        doc.build(story)
        buffer.seek(0)
        return StreamingResponse(
            buffer, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{a["reference"]}.pdf"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- MOUVEMENTS DE STOCK (traçabilité) ---
@app.get("/mouvements-stock")
def lister_mouvements_stock(
    produit: Optional[str] = None, type_mouvement: Optional[str] = None, limite: int = 200,
    current_user: dict = Depends(get_current_user)
):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        conditions, params = [], []
        if produit:
            conditions.append("produit_nom LIKE ?")
            params.append(f"%{produit}%")
        if type_mouvement:
            conditions.append("type_mouvement = ?")
            params.append(type_mouvement)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        params.append(min(limite, 500))
        cur.execute(f"""
            SELECT id, date_mouvement, produit_nom, type_mouvement, quantite, stock_avant, stock_apres,
                   operateur, reference_doc, motif, vente_id, fournisseur_nom
            FROM mouvements_stock {where} ORDER BY id DESC LIMIT ?
        """, params)
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- JOURNAL D'AUDIT (admin uniquement) ---
@app.get("/journal")
def lister_journal(limite: int = 200, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT j.id, j.date_action, j.action, j.detail, u.identifiant as utilisateur
            FROM journal j LEFT JOIN utilisateurs u ON u.id = j.utilisateur_id
            ORDER BY j.id DESC LIMIT ?
        """, (min(limite, 500),))
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- GESTION DES UTILISATEURS (admin uniquement) ---
@app.get("/utilisateurs")
def lister_utilisateurs(current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, identifiant, role, actif, created_at FROM utilisateurs ORDER BY identifiant ASC")
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/utilisateurs", status_code=status.HTTP_201_CREATED)
def creer_utilisateur(data: UtilisateurCreateSchema, current_user: dict = Depends(get_current_admin)):
    identifiant = data.identifiant.strip()
    if not identifiant:
        raise HTTPException(status_code=400, detail="L'identifiant est obligatoire.")
    if len(data.mot_de_passe) < 4:
        raise HTTPException(status_code=400, detail="Le mot de passe doit faire au moins 4 caractères.")
    if data.role not in ROLES_VALIDES:
        raise HTTPException(status_code=400, detail=f"Rôle invalide — choisir parmi : {', '.join(ROLES_VALIDES)}.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id FROM utilisateurs WHERE identifiant = ?", (identifiant,))
        if cur.fetchone():
            conn.close()
            raise HTTPException(status_code=400, detail=f"L'identifiant '{identifiant}' existe déjà.")
        hash_mdp = config.hacher_mdp(data.mot_de_passe)
        cur.execute("INSERT INTO utilisateurs (identifiant, mot_de_passe, role, actif) VALUES (?, ?, ?, 1)",
                    (identifiant, hash_mdp, data.role))
        conn.commit()
        user_id = cur.lastrowid
        conn.close()
        _journaliser(current_user["username"], "UTILISATEUR_CREE", f"{identifiant} ({data.role})")
        return {"status": "success", "id": user_id}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/utilisateurs/{user_id}")
def modifier_utilisateur(user_id: int, data: UtilisateurUpdateSchema, current_user: dict = Depends(get_current_admin)):
    if data.role not in ROLES_VALIDES:
        raise HTTPException(status_code=400, detail=f"Rôle invalide — choisir parmi : {', '.join(ROLES_VALIDES)}.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT identifiant, role, actif FROM utilisateurs WHERE id = ?", (user_id,))
        cible = cur.fetchone()
        if not cible:
            conn.close()
            raise HTTPException(status_code=404, detail="Utilisateur introuvable.")

        # Jamais se désactiver / se rétrograder soi-même — pour éviter un
        # verrouillage accidentel (demander à un autre admin de le faire).
        if cible["identifiant"] == current_user["username"] and (not data.actif or data.role != "admin"):
            conn.close()
            raise HTTPException(status_code=400, detail="Tu ne peux pas modifier ton propre statut ou rôle admin.")

        # Ne jamais se retrouver sans aucun admin actif.
        devient_non_admin_actif = (cible["role"] == "admin") and (data.role != "admin" or not data.actif)
        if devient_non_admin_actif:
            cur.execute("SELECT COUNT(*) FROM utilisateurs WHERE role='admin' AND actif=1 AND id != ?", (user_id,))
            if cur.fetchone()[0] == 0:
                conn.close()
                raise HTTPException(status_code=400, detail="Impossible — ce serait le dernier compte administrateur actif.")

        cur.execute("UPDATE utilisateurs SET role=?, actif=? WHERE id=?", (data.role, 1 if data.actif else 0, user_id))
        conn.commit()
        conn.close()
        _journaliser(current_user["username"], "UTILISATEUR_MODIFIE", f"{cible['identifiant']} → rôle={data.role}, actif={data.actif}")
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/utilisateurs/{user_id}/reset-password")
def reinitialiser_mot_de_passe(user_id: int, data: ResetMotDePasseSchema, current_user: dict = Depends(get_current_admin)):
    if len(data.nouveau_mot_de_passe) < 4:
        raise HTTPException(status_code=400, detail="Le mot de passe doit faire au moins 4 caractères.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT identifiant FROM utilisateurs WHERE id = ?", (user_id,))
        cible = cur.fetchone()
        if not cible:
            conn.close()
            raise HTTPException(status_code=404, detail="Utilisateur introuvable.")
        hash_mdp = config.hacher_mdp(data.nouveau_mot_de_passe)
        cur.execute("UPDATE utilisateurs SET mot_de_passe=? WHERE id=?", (hash_mdp, user_id))
        conn.commit()
        conn.close()
        _journaliser(current_user["username"], "MOT_DE_PASSE_REINITIALISE", cible["identifiant"])
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- COMMANDES FOURNISSEURS (admin uniquement, comme Produits/Fournisseurs) ---
def _commande_to_dict(c):
    total = c["total"] or 0.0
    montant_paye = c["montant_paye"] or 0.0
    return {
        "id": c["id"], "fournisseur_id": c["fournisseur_id"], "fournisseur_nom": c["fournisseur_nom"],
        "statut": c["statut"], "total": total, "montant_paye": montant_paye, "reste_a_payer": max(0.0, total - montant_paye),
        "note": c["note"], "date_commande": c["date_commande"], "date_livraison": c["date_livraison"], "date_reception": c["date_reception"],
    }

@app.get("/commandes-fournisseurs")
def lister_commandes(statut: Optional[str] = None, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        if statut:
            cur.execute("""
                SELECT cf.*, f.nom as fournisseur_nom FROM commandes_fournisseurs cf
                LEFT JOIN fournisseurs f ON f.id = cf.fournisseur_id
                WHERE cf.statut = ? ORDER BY cf.id DESC
            """, (statut,))
        else:
            cur.execute("""
                SELECT cf.*, f.nom as fournisseur_nom FROM commandes_fournisseurs cf
                LEFT JOIN fournisseurs f ON f.id = cf.fournisseur_id
                ORDER BY cf.id DESC
            """)
        rows = [_commande_to_dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/commandes-fournisseurs", status_code=status.HTTP_201_CREATED)
def creer_commande(data: CommandeCreateSchema, current_user: dict = Depends(get_current_admin)):
    if not data.lignes:
        raise HTTPException(status_code=400, detail="La commande doit contenir au moins une ligne.")
    for l in data.lignes:
        if l.quantite <= 0:
            raise HTTPException(status_code=400, detail="Les quantités doivent être positives.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, nom FROM fournisseurs WHERE id = ?", (data.fournisseur_id,))
        fournisseur = cur.fetchone()
        if not fournisseur:
            conn.close()
            raise HTTPException(status_code=404, detail="Fournisseur introuvable.")

        cur.execute("SELECT id FROM utilisateurs WHERE identifiant = ?", (current_user["username"],))
        u = cur.fetchone()
        utilisateur_id = u["id"] if u else None

        total = sum(l.quantite * l.prix_unitaire for l in data.lignes)
        cur.execute("""
            INSERT INTO commandes_fournisseurs (fournisseur_id, utilisateur_id, statut, total, note, date_livraison)
            VALUES (?, ?, 'en_cours', ?, ?, ?)
        """, (data.fournisseur_id, utilisateur_id, total, data.note, data.date_livraison))
        commande_id = cur.lastrowid
        for l in data.lignes:
            cur.execute("""
                INSERT INTO lignes_commande (commande_id, produit_id, quantite, prix_unitaire, quantite_recue)
                VALUES (?, ?, ?, ?, 0)
            """, (commande_id, l.produit_id, l.quantite, l.prix_unitaire))
        conn.commit()
        conn.close()
        _journaliser(current_user["username"], "COMMANDE_CREEE", f"Commande #{commande_id} — {fournisseur['nom']} — {total:.0f} F")
        return {"status": "success", "id": commande_id, "total": total}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/commandes-fournisseurs/{commande_id}")
def detail_commande(commande_id: int, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT cf.*, f.nom as fournisseur_nom FROM commandes_fournisseurs cf
            LEFT JOIN fournisseurs f ON f.id = cf.fournisseur_id WHERE cf.id = ?
        """, (commande_id,))
        c = cur.fetchone()
        if not c:
            conn.close()
            raise HTTPException(status_code=404, detail="Commande introuvable.")
        cur.execute("""
            SELECT lc.id, lc.produit_id, lc.quantite, lc.prix_unitaire, lc.quantite_recue, p.nom, p.unite
            FROM lignes_commande lc JOIN produits p ON p.id = lc.produit_id
            WHERE lc.commande_id = ? ORDER BY p.nom ASC
        """, (commande_id,))
        lignes = [{
            "id": r["id"], "produit_id": r["produit_id"], "produit_nom": r["nom"], "unite": r["unite"],
            "quantite": r["quantite"], "prix_unitaire": r["prix_unitaire"], "quantite_recue": r["quantite_recue"],
            "reste_a_recevoir": max(0.0, r["quantite"] - r["quantite_recue"]),
        } for r in cur.fetchall()]
        conn.close()
        return {"commande": _commande_to_dict(c), "lignes": lignes}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/commandes-fournisseurs/{commande_id}/reception")
def reception_commande(commande_id: int, data: ReceptionCommandeSchema, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, statut, fournisseur_id FROM commandes_fournisseurs WHERE id = ?", (commande_id,))
        commande = cur.fetchone()
        if not commande:
            conn.close()
            raise HTTPException(status_code=404, detail="Commande introuvable.")
        if commande["statut"] in ("annulee", "receptionnee"):
            conn.close()
            raise HTTPException(status_code=400, detail="Cette commande est déjà clôturée (reçue ou annulée).")

        cur.execute("SELECT nom FROM fournisseurs WHERE id = ?", (commande["fournisseur_id"],))
        f = cur.fetchone()
        fournisseur_nom = f["nom"] if f else ""

        for ligne_data in data.lignes:
            if ligne_data.quantite_recue <= 0:
                continue
            cur.execute("SELECT produit_id, quantite, quantite_recue, prix_unitaire FROM lignes_commande WHERE id = ? AND commande_id = ?",
                        (ligne_data.ligne_id, commande_id))
            ligne = cur.fetchone()
            if not ligne:
                continue
            reste = ligne["quantite"] - ligne["quantite_recue"]
            if ligne_data.quantite_recue > reste + 1e-9:
                conn.close()
                raise HTTPException(status_code=400, detail=f"Quantité reçue ({ligne_data.quantite_recue}) supérieure au reste à recevoir ({reste}) pour une ligne.")

            cur.execute("UPDATE lignes_commande SET quantite_recue = quantite_recue + ? WHERE id = ?",
                        (ligne_data.quantite_recue, ligne_data.ligne_id))
            database.enregistrer_mouvement(
                conn, ligne["produit_id"], "ENTREE_COMMANDE", ligne_data.quantite_recue,
                current_user["username"], reference_doc=f"COMMANDE-{commande_id}",
                fournisseur_id=commande["fournisseur_id"], fournisseur_nom=fournisseur_nom,
                prix_unitaire=ligne["prix_unitaire"]
            )

        # Statut global : tout reçu -> receptionnee ; sinon partielle
        cur.execute("SELECT SUM(quantite), SUM(quantite_recue) FROM lignes_commande WHERE commande_id = ?", (commande_id,))
        total_commande, total_recu = cur.fetchone()
        nouveau_statut = "receptionnee" if total_recu >= total_commande - 1e-9 else "partielle"
        cur.execute("""
            UPDATE commandes_fournisseurs SET statut=?,
                date_reception = COALESCE(date_reception, datetime('now','localtime'))
            WHERE id=?
        """, (nouveau_statut, commande_id))

        conn.commit()
        conn.close()
        _journaliser(current_user["username"], "COMMANDE_RECEPTIONNEE", f"Commande #{commande_id} — statut={nouveau_statut}")
        return {"status": "success", "nouveau_statut": nouveau_statut}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/commandes-fournisseurs/{commande_id}/paiement")
def paiement_commande(commande_id: int, data: PaiementCommandeSchema, current_user: dict = Depends(get_current_admin)):
    if data.montant <= 0:
        raise HTTPException(status_code=400, detail="Le montant doit être positif.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT total, montant_paye FROM commandes_fournisseurs WHERE id = ?", (commande_id,))
        c = cur.fetchone()
        if not c:
            conn.close()
            raise HTTPException(status_code=404, detail="Commande introuvable.")
        reste = c["total"] - c["montant_paye"]
        if data.montant > reste + 1e-9:
            conn.close()
            raise HTTPException(status_code=400, detail=f"Le montant ({data.montant:.0f} F) dépasse le reste à payer ({reste:.0f} F).")
        cur.execute("UPDATE commandes_fournisseurs SET montant_paye = montant_paye + ? WHERE id = ?", (data.montant, commande_id))
        conn.commit()
        conn.close()
        _journaliser(current_user["username"], "COMMANDE_PAIEMENT", f"Commande #{commande_id} — {data.montant:.0f} F")
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/commandes-fournisseurs/{commande_id}/annuler")
def annuler_commande(commande_id: int, current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT statut, montant_paye FROM commandes_fournisseurs WHERE id = ?", (commande_id,))
        c = cur.fetchone()
        if not c:
            conn.close()
            raise HTTPException(status_code=404, detail="Commande introuvable.")
        if c["statut"] in ("receptionnee", "partielle"):
            conn.close()
            raise HTTPException(status_code=400, detail="Impossible d'annuler une commande déjà (partiellement) reçue.")
        if c["montant_paye"] > 0:
            conn.close()
            raise HTTPException(status_code=400, detail="Impossible d'annuler une commande déjà payée en partie.")
        cur.execute("UPDATE commandes_fournisseurs SET statut='annulee' WHERE id = ?", (commande_id,))
        conn.commit()
        conn.close()
        _journaliser(current_user["username"], "COMMANDE_ANNULEE", f"Commande #{commande_id}")
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- CLIENTS ---
@app.get("/clients")
def lister_clients(q: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        if q:
            cur.execute("""SELECT id, nom, telephone, email, adresse, encours_max, solde_credit
                           FROM clients WHERE nom LIKE ? OR telephone LIKE ? ORDER BY nom ASC""",
                        (f"%{q}%", f"%{q}%"))
        else:
            cur.execute("SELECT id, nom, telephone, email, adresse, encours_max, solde_credit FROM clients ORDER BY nom ASC")
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/clients", status_code=status.HTTP_201_CREATED)
def creer_client(data: ClientSchema, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""INSERT INTO clients (nom, telephone, email, adresse, encours_max, solde_credit)
                       VALUES (?, ?, ?, ?, ?, 0)""",
                    (data.nom, data.telephone, data.email, data.adresse, data.encours_max))
        conn.commit()
        client_id = cur.lastrowid
        conn.close()
        return {"status": "success", "id": client_id}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/clients/{client_id}")
def modifier_client(client_id: int, data: ClientSchema, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""UPDATE clients SET nom=?, telephone=?, email=?, adresse=?, encours_max=? WHERE id=?""",
                    (data.nom, data.telephone, data.email, data.adresse, data.encours_max, client_id))
        if cur.rowcount == 0:
            conn.close()
            raise HTTPException(status_code=404, detail="Client introuvable.")
        conn.commit()
        conn.close()
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- HISTORIQUE D'UN CLIENT (ventes à crédit + règlements, fusionnés et triés) ---
@app.get("/clients/{client_id}/historique")
def historique_client(client_id: int, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, nom, telephone, email, adresse, encours_max, solde_credit FROM clients WHERE id=?", (client_id,))
        client = cur.fetchone()
        if not client:
            conn.close()
            raise HTTPException(status_code=404, detail="Client introuvable.")

        cur.execute("""SELECT id, date_vente, total, recu FROM ventes
                       WHERE client_id=? AND mode_paiement='CREDIT' AND annulee=0 ORDER BY date_vente DESC""", (client_id,))
        ventes = [{"type": "vente", "date": r["date_vente"], "vente_id": r["id"],
                    "montant": r["total"] - r["recu"]} for r in cur.fetchall()]

        cur.execute("""SELECT id, montant, mode_paiement, note, date_reglement FROM reglements_credit
                       WHERE client_id=? ORDER BY date_reglement DESC""", (client_id,))
        reglements = [{"type": "reglement", "date": r["date_reglement"], "reglement_id": r["id"],
                        "montant": -r["montant"], "mode_paiement": r["mode_paiement"], "note": r["note"]} for r in cur.fetchall()]

        historique = sorted(ventes + reglements, key=lambda x: x["date"], reverse=True)
        conn.close()
        return {"client": dict(client), "historique": historique}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- ENREGISTRER UN REGLEMENT (paiement reçu d'un client) ---
@app.post("/clients/{client_id}/reglement", status_code=status.HTTP_201_CREATED)
def enregistrer_reglement(client_id: int, data: ReglementSchema, current_user: dict = Depends(get_current_user)):
    if data.montant <= 0:
        raise HTTPException(status_code=400, detail="Le montant doit être positif.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT nom, solde_credit FROM clients WHERE id=?", (client_id,))
        client = cur.fetchone()
        if not client:
            conn.close()
            raise HTTPException(status_code=404, detail="Client introuvable.")
        if data.montant > client["solde_credit"]:
            conn.close()
            raise HTTPException(
                status_code=400,
                detail=f"Le montant ({data.montant:.0f} F) dépasse le solde dû par {client['nom']} ({client['solde_credit']:.0f} F)."
            )
        note_complete = f"[{current_user['username']}] {data.note}".strip()
        cur.execute("""INSERT INTO reglements_credit (client_id, montant, mode_paiement, note)
                       VALUES (?, ?, ?, ?)""", (client_id, data.montant, data.mode_paiement, note_complete))
        cur.execute("UPDATE clients SET solde_credit = solde_credit - ? WHERE id=?", (data.montant, client_id))
        conn.commit()
        conn.close()
        _journaliser(current_user["username"], "REGLEMENT_CLIENT", f"{client['nom']} — {data.montant:.0f} F ({data.mode_paiement})")
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/ventes", status_code=status.HTTP_201_CREATED)
def creer_une_vente(vente_data: VenteCreateSchema, current_user: dict = Depends(get_current_user)):
    if not vente_data.lignes:
        raise HTTPException(status_code=400, detail="Le panier est vide.")
    try:
        conn = database.get_connection()
        cur = conn.cursor()

        # --- Vérification du stock AVANT toute écriture ---
        # On lit le stock courant de chaque produit demandé et on refuse la
        # vente en bloc si une ligne dépasse le disponible (pas de vente
        # partielle silencieuse).
        manques = []
        noms_produits = {}
        for ligne in vente_data.lignes:
            cur.execute("SELECT nom, stock_actuel FROM produits WHERE id = ?", (ligne.produit_id,))
            produit = cur.fetchone()
            if not produit:
                conn.close()
                raise HTTPException(status_code=404, detail=f"Produit id={ligne.produit_id} introuvable.")
            noms_produits[ligne.produit_id] = produit["nom"]
            disponible = produit["stock_actuel"] or 0
            if ligne.quantite > disponible:
                manques.append(f"{produit['nom']} (demandé : {ligne.quantite}, disponible : {disponible})")

        if manques:
            conn.close()
            raise HTTPException(status_code=400, detail="Stock insuffisant — " + " ; ".join(manques))

        # --- Vérification du client pour une vente à crédit ---
        client = None
        if vente_data.mode_paiement == "CREDIT":
            if not vente_data.client_id:
                conn.close()
                raise HTTPException(status_code=400, detail="Un client est obligatoire pour une vente à crédit.")
            cur.execute("SELECT id, nom, solde_credit, encours_max FROM clients WHERE id = ?", (vente_data.client_id,))
            client = cur.fetchone()
            if not client:
                conn.close()
                raise HTTPException(status_code=404, detail="Client introuvable.")

        total_brut = sum(l.quantite * l.prix_unitaire for l in vente_data.lignes)
        total_net = max(0.0, total_brut - vente_data.remise)

        # Pour une vente CREDIT, "recu" est l'acompte versé (0 par défaut =
        # entièrement à crédit) — contrairement aux autres modes où l'absence
        # de "recu" explicite vaut paiement intégral.
        if vente_data.mode_paiement == "CREDIT":
            recu = max(0.0, vente_data.recu)
        else:
            recu = vente_data.recu if vente_data.recu > 0 else total_net
        monnaie = max(0.0, recu - total_net)

        # Plafond de crédit (encours_max=0 → pas de limite)
        montant_credit = 0.0
        if vente_data.mode_paiement == "CREDIT":
            montant_credit = max(0.0, total_net - recu)
            nouveau_solde = client["solde_credit"] + montant_credit
            if client["encours_max"] > 0 and nouveau_solde > client["encours_max"]:
                conn.close()
                raise HTTPException(
                    status_code=400,
                    detail=f"Plafond de crédit dépassé pour {client['nom']} — "
                           f"solde actuel : {client['solde_credit']:.0f} F, plafond : {client['encours_max']:.0f} F, "
                           f"cette vente ajouterait : {montant_credit:.0f} F."
                )

        cur.execute("""
            INSERT INTO ventes (total, recu, monnaie, vendeur, mode_paiement, client_id, remise, montant_regle, note, cloture, annulee)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0)
        """, (total_net, recu, monnaie, current_user["username"], vente_data.mode_paiement, vente_data.client_id, vente_data.remise, recu, vente_data.note))

        vente_id = cur.lastrowid
        for ligne in vente_data.lignes:
            cur.execute("INSERT INTO lignes_vente (vente_id, produit_id, quantite, prix_unitaire, remise) VALUES (?, ?, ?, ?, ?)",
                        (vente_id, ligne.produit_id, ligne.quantite, ligne.prix_unitaire, ligne.remise))
            # Passe par database.py : décrémente stock ET stock_actuel ensemble,
            # et journalise dans mouvements_stock (comme le fait le desktop).
            database.enregistrer_mouvement(
                conn, ligne.produit_id, "SORTIE_VENTE", ligne.quantite,
                current_user["username"], vente_id=vente_id,
                prix_unitaire=ligne.prix_unitaire, reference_doc=f"VENTE-WEB-{vente_id}"
            )

        # Mise à jour du solde client si vente à crédit
        if vente_data.mode_paiement == "CREDIT" and montant_credit > 0:
            cur.execute("UPDATE clients SET solde_credit = solde_credit + ? WHERE id = ?",
                        (montant_credit, vente_data.client_id))

        conn.commit()

        # --- Émission FEC ---
        # On ne bloque jamais la vente si la FEC échoue (la caisse doit
        # continuer à fonctionner) mais on journalise l'erreur clairement :
        # une vente "fec": null doit être régularisée manuellement ensuite.
        fec_info = None
        try:
            ifu, _rccm = config.get_infos_fiscales()
            f = fec.emettre_facture_fec(conn, vente_id, total_net, ifu)
            fec_info = {
                "numero": f["numero"],
                "date_heure": f["date_heure"],
                "signature": f["signature"],
                "qr_base64": base64.b64encode(f["qr_bytes"]).decode("ascii") if f.get("qr_bytes") else None,
            }
        except Exception as e:
            config.logger.error(f"FEC: échec d'émission pour vente_id={vente_id} — {e}", exc_info=True)

        # --- Impression directe (ESC/POS) ---
        # N'imprime que si l'imprimante est activée dans Paramètres > Imprimante.
        # Ne bloque jamais la vente : une panne d'imprimante ne doit jamais
        # empêcher l'enregistrement d'une vente (impression.py gère déjà ses
        # propres erreurs en interne et journalise dans hshop_v2.log).
        try:
            panier_impression = [
                {"nom": noms_produits.get(l.produit_id, f"Produit #{l.produit_id}"),
                 "qty": l.quantite, "prix": l.prix_unitaire, "remise": l.remise}
                for l in vente_data.lignes
            ]
            impression.imprimer_ticket_thermique(
                vendeur=current_user["username"], panier=panier_impression,
                total=total_net, recu=recu, monnaie=monnaie, id_vente=vente_id,
                mode=vente_data.mode_paiement, client_nom=(client["nom"] if client else None),
                remise_globale=vente_data.remise,
            )
        except Exception as e:
            config.logger.error(f"Impression directe: échec pour vente_id={vente_id} — {e}", exc_info=True)

        conn.close()
        _journaliser(current_user["username"], "VENTE", f"Vente #{vente_id} — {total_net:.0f} F — {vente_data.mode_paiement}")
        return {"status": "success", "vente_id": vente_id, "fec": fec_info}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- ROUTE 4 : HISTORIQUE DES 30 DERNIERES VENTES ---
@app.get("/ventes/historique")
def recuperer_historique_ventes(current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, date_vente, total, mode_paiement, vendeur, note FROM ventes ORDER BY id DESC LIMIT 30")
        ventes = cur.fetchall()
        resultat = [{"id": v["id"], "date_vente": v["date_vente"], "total": v["total"], "mode_paiement": v["mode_paiement"], "vendeur": v["vendeur"], "note": v["note"]} for v in ventes]
        conn.close()
        return resultat
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- ROUTE 5 : DETAIL D'UNE VENTE PRECISE ---
@app.get("/ventes-details/{vente_id}")
def recuperer_details_vente(vente_id: int, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT lv.quantite, lv.prix_unitaire, lv.remise, p.nom as produit_nom, p.unite
            FROM lignes_vente lv
            JOIN produits p ON lv.produit_id = p.id
            WHERE lv.vente_id = ?
        """, (vente_id,))
        lignes = cur.fetchall()
        conn.close()
        return [{"produit_nom": l["produit_nom"], "quantite": l["quantite"], "prix_unitaire": l["prix_unitaire"], "remise": l["remise"], "unite": l["unite"] if l["unite"] else "pce"} for l in lignes]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- ROUTE 6 : APPROVISIONNER / AJOUTER DU STOCK ---
@app.post("/produits/{produit_id}/approvisionner")
def approvisionner_produit(produit_id: int, data: ApprovisionnerSchema, current_user: dict = Depends(get_current_user)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()

        cur.execute("SELECT nom FROM produits WHERE id = ?", (produit_id,))
        produit = cur.fetchone()
        if not produit:
            conn.close()
            raise HTTPException(status_code=404, detail="Produit non trouvé.")

        # Synchro stock + stock_actuel + journal mouvements_stock, comme le desktop.
        database.enregistrer_mouvement(
            conn, produit_id, "ENTREE_MANUELLE", data.quantite,
            current_user["username"], reference_doc="Appro Web"
        )
        conn.commit()
        conn.close()
        return {"status": "success", "message": f"Stock mis à jour pour le produit {produit_id}"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/dashboard/stats")
def dashboard_stats(current_user: dict = Depends(get_current_admin)):
    try:
        conn = database.get_connection()
        cur = conn.cursor()

        # CA jour / hier (pour la variation %)
        cur.execute("SELECT COALESCE(SUM(total),0) FROM ventes WHERE date(date_vente)=date('now','localtime') AND annulee=0")
        ca_jour = cur.fetchone()[0]
        cur.execute("SELECT COALESCE(SUM(total),0) FROM ventes WHERE date(date_vente)=date('now','localtime','-1 day') AND annulee=0")
        ca_hier = cur.fetchone()[0]
        variation_pct = round((ca_jour - ca_hier) / ca_hier * 100, 1) if ca_hier > 0 else None

        # CA mois + panier moyen
        cur.execute("""
            SELECT COALESCE(SUM(total),0), COUNT(*)
            FROM ventes
            WHERE strftime('%Y-%m', date_vente) = strftime('%Y-%m','now','localtime') AND annulee=0
        """)
        ca_mois, nb_ventes_mois = cur.fetchone()
        panier_moyen_mois = round(ca_mois / nb_ventes_mois, 0) if nb_ventes_mois else 0

        # Net caisse du jour : encaissé hors crédit (le crédit n'est pas
        # encaissé tant qu'il n'est pas réglé) moins les dépenses du jour.
        cur.execute("""
            SELECT COALESCE(SUM(total),0) FROM ventes
            WHERE mode_paiement != 'CREDIT' AND date(date_vente)=date('now','localtime') AND annulee=0
        """)
        encaisse_jour = cur.fetchone()[0]
        cur.execute("SELECT COALESCE(SUM(montant),0) FROM depenses WHERE date(date_depense)=date('now','localtime')")
        depenses_jour = cur.fetchone()[0]
        net_caisse_jour = encaisse_jour - depenses_jour

        # Répartition des modes de paiement (30 derniers jours)
        cur.execute("""
            SELECT mode_paiement, COUNT(*) nb, COALESCE(SUM(total),0) total
            FROM ventes
            WHERE annulee=0 AND date(date_vente) >= date('now','localtime','-30 days')
            GROUP BY mode_paiement
            ORDER BY total DESC
        """)
        modes_paiement = [{"mode": r["mode_paiement"], "nb": r["nb"], "total": r["total"]} for r in cur.fetchall()]

        # Top 5 produits du jour — repli sur 7 jours si rien n'a encore été vendu aujourd'hui
        cur.execute("""
            SELECT p.nom, SUM(lv.quantite) qte, SUM(lv.quantite*lv.prix_unitaire) total
            FROM lignes_vente lv
            JOIN ventes v ON v.id = lv.vente_id
            JOIN produits p ON p.id = lv.produit_id
            WHERE date(v.date_vente) = date('now','localtime') AND v.annulee = 0
            GROUP BY p.id ORDER BY total DESC LIMIT 5
        """)
        rows = cur.fetchall()
        periode_top = "aujourd'hui"
        if not rows:
            cur.execute("""
                SELECT p.nom, SUM(lv.quantite) qte, SUM(lv.quantite*lv.prix_unitaire) total
                FROM lignes_vente lv
                JOIN ventes v ON v.id = lv.vente_id
                JOIN produits p ON p.id = lv.produit_id
                WHERE date(v.date_vente) >= date('now','localtime','-6 days') AND v.annulee = 0
                GROUP BY p.id ORDER BY total DESC LIMIT 5
            """)
            rows = cur.fetchall()
            periode_top = "7 derniers jours"
        top_produits = [{"nom": r["nom"], "quantite": r["qte"], "total": r["total"]} for r in rows]

        # Dernières ventes
        cur.execute("""
            SELECT id, date_vente, total, mode_paiement, vendeur
            FROM ventes WHERE annulee=0 ORDER BY id DESC LIMIT 10
        """)
        dernieres_ventes = [{"id": r["id"], "date_vente": r["date_vente"], "total": r["total"],
                              "mode_paiement": r["mode_paiement"], "vendeur": r["vendeur"]} for r in cur.fetchall()]

        # Graphique 7 jours (calculé côté Python pour ne pas dépendre du fuseau SQLite)
        cur.execute("""
            SELECT date(date_vente) jour, COALESCE(SUM(total),0) ca
            FROM ventes WHERE annulee=0 AND date(date_vente) >= date('now','localtime','-6 days')
            GROUP BY jour ORDER BY jour ASC
        """)
        ca_par_jour = {r["jour"]: r["ca"] for r in cur.fetchall()}
        jours_7 = [(datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6, -1, -1)]
        graphique_7j = [{"jour": j, "ca": ca_par_jour.get(j, 0)} for j in jours_7]

        # Alertes stock bas (seuil produit si défini, sinon seuil global config)
        seuil_defaut = config.get_seuil_stock()
        cur.execute("""
            SELECT id, nom, stock_actuel, stock_min, unite
            FROM produits
            WHERE actif=1 AND stock_actuel <= CASE WHEN stock_min > 0 THEN stock_min ELSE ? END
            ORDER BY stock_actuel ASC LIMIT 15
        """, (seuil_defaut,))
        stock_bas = [{"id": r["id"], "nom": r["nom"], "stock": r["stock_actuel"], "stock_min": r["stock_min"], "unite": r["unite"]} for r in cur.fetchall()]

        # Crédits clients impayés
        cur.execute("SELECT id, nom, solde_credit, telephone FROM clients WHERE solde_credit > 0 ORDER BY solde_credit DESC LIMIT 15")
        credits_impayes = [{"id": r["id"], "nom": r["nom"], "solde": r["solde_credit"], "telephone": r["telephone"]} for r in cur.fetchall()]
        cur.execute("SELECT COALESCE(SUM(solde_credit),0) FROM clients WHERE solde_credit > 0")
        total_impaye = cur.fetchone()[0]

        conn.close()
        return {
            "ca_jour": ca_jour, "ca_hier": ca_hier, "variation_pct": variation_pct,
            "ca_mois": ca_mois, "nb_ventes_mois": nb_ventes_mois, "panier_moyen_mois": panier_moyen_mois,
            "net_caisse_jour": net_caisse_jour, "depenses_jour": depenses_jour,
            "modes_paiement": modes_paiement,
            "top_produits": top_produits, "periode_top": periode_top,
            "dernieres_ventes": dernieres_ventes,
            "graphique_7j": graphique_7j,
            "stock_bas": stock_bas,
            "credits_impayes": credits_impayes, "total_impaye": total_impaye,
        }
    except Exception as e:
        if 'conn' in locals(): conn.close()
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import sys
    sys.stdout.flush()
    # 127.0.0.1 : accessible uniquement depuis ce PC. Le jour où tu veux
    # plusieurs postes (Tailscale ou autre réseau), remplace par "0.0.0.0"
    # — c'est la seule ligne à changer, le reste (URL relative dans le JS,
    # CORS same-origin) fonctionne déjà sans modification.
    uvicorn.run(app, host="0.0.0.0", port=8080)
